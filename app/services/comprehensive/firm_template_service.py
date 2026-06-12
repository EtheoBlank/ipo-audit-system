"""多所模板管理 + 历史底稿库服务。

提供：
- FirmTemplateService：模板上传、列表、按 firm 隔离、版本管理
- HistoricalLibraryService：历史底稿脱敏、入库、检索（作为第 4 类信息源）
"""
from __future__ import annotations

import hashlib
import logging
import re
from io import BytesIO
from typing import Iterable, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.db_models import FirmTemplate, HistoricalWorkpaper
from app.services.comprehensive.schemas import TemplateSchema
from app.services.comprehensive.template_parser import TemplateParser
from app.services.comprehensive.web_search_engine import SearchHit

logger = logging.getLogger(__name__)


# ============================== 多所模板服务 ==============================

class FirmTemplateService:
    """多所模板管理。"""

    def __init__(self, session: AsyncSession):
        self.session = session

    async def upload(
        self,
        firm_id: str,
        template_id: str,
        version: str,
        template_bytes: bytes,
        template_name: str,
        *,
        industry: Optional[str] = None,
        audit_period: Optional[str] = None,
        description: Optional[str] = None,
        created_by: Optional[str] = None,
    ) -> FirmTemplate:
        """上传一份模板，自动解析后存档。"""
        # 解析一次以生成 field_schema_json 快照
        parser = TemplateParser()
        try:
            schema = parser.parse(template_bytes)
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"模板解析失败: {exc}") from exc

        # 如果同 (firm_id, template_id, version) 已存在则覆盖
        existing = await self._find(firm_id, template_id, version)
        if existing is not None:
            existing.template_bytes = template_bytes
            existing.template_name = template_name
            existing.industry = industry or schema.industry
            existing.audit_period = audit_period or schema.audit_period
            existing.field_schema_json = schema.model_dump_json()
            existing.description = description
            existing.is_active = True
            await self.session.commit()
            return existing

        t = FirmTemplate(
            firm_id=firm_id,
            template_id=template_id,
            version=version,
            template_name=template_name,
            industry=industry or schema.industry,
            audit_period=audit_period or schema.audit_period,
            template_bytes=template_bytes,
            field_schema_json=schema.model_dump_json(),
            description=description,
            created_by=created_by,
        )
        self.session.add(t)
        await self.session.commit()
        await self.session.refresh(t)
        return t

    async def get_latest(
        self, firm_id: str, template_id: str, version: Optional[str] = None
    ) -> Optional[FirmTemplate]:
        """获取最新（或指定版本）的模板。"""
        if version is not None:
            return await self._find(firm_id, template_id, version)
        # 取 published_at 最新的
        stmt = (
            select(FirmTemplate)
            .where(
                FirmTemplate.firm_id == firm_id,
                FirmTemplate.template_id == template_id,
                FirmTemplate.is_active.is_(True),
            )
            .order_by(FirmTemplate.published_at.desc())
            .limit(1)
        )
        return (await self.session.execute(stmt)).scalars().first()

    async def list_for_firm(self, firm_id: str) -> list[FirmTemplate]:
        """列出某事务所的所有模板。"""
        stmt = (
            select(FirmTemplate)
            .where(FirmTemplate.firm_id == firm_id, FirmTemplate.is_active.is_(True))
            .order_by(FirmTemplate.template_id, FirmTemplate.published_at.desc())
        )
        return list((await self.session.execute(stmt)).scalars().all())

    async def deactivate(self, firm_id: str, template_id: str, version: str) -> bool:
        """软删除：把 is_active 置为 False。"""
        t = await self._find(firm_id, template_id, version)
        if t is None:
            return False
        t.is_active = False
        await self.session.commit()
        return True

    async def parse_to_schema(self, firm_id: str, template_id: str) -> Optional[TemplateSchema]:
        """从库中读出模板并解析为 TemplateSchema。

        优先使用 ``field_schema_json`` 快照（避免每次重新解析 .xlsx），
        快照不存在或解析失败时回退到重新解析。
        """
        t = await self.get_latest(firm_id, template_id)
        if t is None:
            return None
        if t.field_schema_json:
            try:
                return TemplateSchema.model_validate_json(t.field_schema_json)
            except Exception:  # noqa: BLE001
                logger.warning("快照反序列化失败，回退到重新解析")
        return TemplateParser().parse(t.template_bytes)

    async def _find(self, firm_id: str, template_id: str, version: str) -> Optional[FirmTemplate]:
        stmt = select(FirmTemplate).where(
            FirmTemplate.firm_id == firm_id,
            FirmTemplate.template_id == template_id,
            FirmTemplate.version == version,
        )
        return (await self.session.execute(stmt)).scalars().first()


# ============================== 历史底稿库 ==============================

# 脱敏规则：覆盖企业实体 / 人名 / 银行账号 / 身份证 / 信用代码 / 合同号等
_ENTITY_PATTERNS = [
    # 中文企业（含有限/股份/集团后缀）
    re.compile(r"[一-龥]{2,15}(?:股份有限公司|有限责任公司|有限公司)"),
    re.compile(r"[一-龥A-Za-z0-9·　]{2,20}(?:集团|控股|总公司)"),
    re.compile(r"[一-龥A-Za-z0-9·　]{2,15}公司"),
    # 中文机构（银行/医院/学校/事务所/厂/局/中心/合作社/基金会）
    re.compile(r"[一-龥]{2,20}(?:银行|医院|学校|事务所|事务所|大学|学院|厂|局|中心|合作社|基金会|研究院|研究所)"),
    # 四大 + 常见咨询 / 律所 / 评级机构
    re.compile(r"\b(?:PWC|KPMG|Deloitte|EY|安永|毕马威|普华永道|德勤|永安|中注协|大华|天健|立信|致同|信永中和|大信|中审众环|容诚|天衡|公证天业|祥恒)\b"),
    # 英文公司（带 Corp/Inc/Ltd/LLC/LLP/Company/Group 等后缀）
    re.compile(r"\b[A-Z][A-Za-z0-9&.\- ]{1,40}\s(?:Corp\.?|Inc\.?|Ltd\.?|LLC|LLP|Company|Co\.,?\sInc|GmbH|AG|S\.A\.|Group|PLC)\b"),
    re.compile(r"\b[A-Z][A-Za-z0-9&.\- ]{1,40}\s(?:Corporation|Company|Group|Holdings|Industries)\b"),
    # 统一社会信用代码（18 位）
    re.compile(r"[0-9A-HJ-NPQRTUWXY]{2}\d{6}[0-9A-HJ-NPQRTUWXY]{10}"),
    # 身份证号（18 位）
    re.compile(r"\b[1-9]\d{5}(?:19|20)\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])\d{3}[\dXx]\b"),
    # 银行卡号（13~19 位连续数字）
    re.compile(r"\b\d{13,19}\b"),
    # 手机号
    re.compile(r"\b1[3-9]\d{9}\b"),
    # 邮箱
    re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"),
    # 合同号 / 函证号 / 项目编号（KB-XYZ-2024-001 等常见形式）
    re.compile(r"\b[A-Z]{1,5}[-_]\d{2,4}[-_]\d{2,6}[-_]\d{2,8}\b"),
]

# 中文姓名（2~4 字，且后接敬称 / 职务 / 项目组 / "先生"/"女士"）
_PERSON_NAME_PATTERNS = [
    # 姓名（独立 2~4 字中文，2~4 字都有）
    re.compile(r"(?<![一-龥])([一-龥]{2,4})(?:先生|女士|同志|律师|教授|博士|医生)(?![一-龥])"),
    re.compile(r"((?:项目负责人|项目经理|签字会计师|审计师|合伙人|质量复核|项目组|复核人)[：:]\s*)?([一-龥]{2,4})"),
]

# 中文姓名模式：2~3 字普通姓名（仅在特定上下文如"审计员 X"中触发，避免误伤正常词）
_PERSON_NAME_CONTEXTUAL = re.compile(
    r"(?:审计员|审计师|项目负责人|项目经理|签字会计师|质量复核人|复核人|合伙人|签字人|编制人|被审计单位(?:的)?(?:法人|董事长|总经理|财务总监|财务负责人)?[：:]?)\s*"
    r"([一-龥]{2,4})"
)


def _anonymize_text(text: str, counter: dict[str, int]) -> str:
    """把 PII 替换为 ``<ENT_n>`` / ``<PER_n>``。"""
    out = text

    # 1) 实体（公司/机构/编号/账号）
    for pat in _ENTITY_PATTERNS:
        out = _replace_pat(out, pat, counter, prefix="ENT")

    # 2) 人名（带敬称/职务上下文）
    for pat in _PERSON_NAME_PATTERNS:
        out = _replace_pat(out, pat, counter, prefix="PER", group=2)
    out = _replace_pat(out, _PERSON_NAME_CONTEXTUAL, counter, prefix="PER", group=1)
    return out


def _replace_pat(text: str, pat: re.Pattern, counter: dict[str, int], prefix: str, group: int | None = None) -> str:
    """用正则匹配并替换为 ``<{prefix}_n>``，相同 token 复用同一编号。"""
    def _sub(m: re.Match) -> str:
        if group is not None and group <= len(m.groups()):
            name = m.group(group)
        else:
            name = m.group(0)
        if not name:
            return m.group(0)
        if name not in counter:
            counter[name] = len(counter) + 1
        return f"<{prefix}_{counter[name]}>"
    return pat.sub(_sub, text)


def _anonymize_excel(b: bytes) -> tuple[bytes, str]:
    """对 xlsx 字节流做脱敏。

    处理范围：
      - 单元格文本（值）
      - 工作表名
      - 单元格批注（comments）
      - 文档元数据（core properties: creator / lastModifiedBy / title）

    返回：(脱敏后字节, 抽取的文本摘要)
    """
    from openpyxl import load_workbook

    counter: dict[str, int] = {}
    wb = load_workbook(filename=BytesIO(b), data_only=False)
    excerpts: list[str] = []

    # 1) 工作表名
    for ws in wb.worksheets:
        if ws.title and any(c in ws.title for c in ("公司", "客户", "供应商", "银行")):
            ws.title = _anonymize_text(ws.title, counter)

    # 2) 单元格值 + 批注
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    anon = _anonymize_text(cell.value, counter)
                    if anon != cell.value:
                        cell.value = anon
                    excerpts.append(anon)
                if cell.comment and cell.comment.text:
                    cell.comment.text = _anonymize_text(cell.comment.text, counter)

    # 3) 文档元数据
    props = wb.properties
    if props:
        if props.creator:
            props.creator = "anonymized"
        if props.lastModifiedBy:
            props.lastModifiedBy = "anonymized"
        if props.title and any(c in props.title for c in ("公司", "客户")):
            props.title = _anonymize_text(props.title, counter)
        if props.subject and any(c in (props.subject or "") for c in ("公司", "客户")):
            props.subject = _anonymize_text(props.subject, counter)

    out_buf = BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue(), "\n".join(excerpts)[:50000]


class HistoricalLibraryService:
    """历史综合底稿库（脱敏 + 入库 + 检索）。"""

    def __init__(self, session: AsyncSession):
        self.session = session

    async def ingest(
        self,
        firm_id: str,
        template_id: str,
        workpaper_bytes: bytes,
        *,
        industry: Optional[str] = None,
        fiscal_year: Optional[int] = None,
        uploaded_by: Optional[str] = None,
    ) -> HistoricalWorkpaper:
        """把一份历史综合底稿脱敏后入库。"""
        anon_bytes, excerpt = _anonymize_excel(workpaper_bytes)
        # 用脱敏前 bytes 计算 hash 作为去重 / 来源标记
        source_hash = hashlib.sha256(workpaper_bytes).hexdigest()[:32]
        rec = HistoricalWorkpaper(
            firm_id=firm_id,
            template_id=template_id,
            project_industry=industry,
            project_fiscal_year=fiscal_year,
            anonymized_bytes=anon_bytes,
            text_excerpt=excerpt,
            source_project_hash=source_hash,
            uploaded_by=uploaded_by,
        )
        self.session.add(rec)
        await self.session.commit()
        await self.session.refresh(rec)
        return rec

    async def search(
        self,
        firm_id: str,
        template_id: str,
        query: str,
        top_k: int = 5,
    ) -> list[SearchHit]:
        """在历史底稿库中按关键词检索，返回 SearchHit 列表。

        实现要点：
          - 关键词过滤在 SQL 端完成（``text_excerpt.contains(kw)`` + OR），
            避免全表加载到 Python
          - 加 ``.limit(top_k * 5)`` 提前截断（再多也用不到）
        """
        from sqlalchemy import or_

        keywords = [w for w in query.split() if len(w) >= 2]
        if not keywords:
            return []

        or_clauses = [
            HistoricalWorkpaper.text_excerpt.contains(kw) for kw in keywords
        ]
        stmt = (
            select(HistoricalWorkpaper)
            .where(
                HistoricalWorkpaper.firm_id == firm_id,
                HistoricalWorkpaper.template_id == template_id,
                or_(*or_clauses),
            )
            .limit(top_k * 5)
        )
        rows = list((await self.session.execute(stmt)).scalars().all())
        if not rows:
            return []

        # 简单打分
        scored: list[tuple[float, HistoricalWorkpaper]] = []
        for r in rows:
            if not r.text_excerpt:
                continue
            score = sum(r.text_excerpt.count(kw) for kw in keywords)
            if score > 0:
                scored.append((float(score), r))
        scored.sort(key=lambda t: t[0], reverse=True)

        hits: list[SearchHit] = []
        for score, r in scored[:top_k]:
            snippet = (r.text_excerpt or "")[:400]
            hits.append(
                SearchHit(
                    title=f"历史底稿 #{r.id} ({r.project_industry or '未知行业'}/{r.project_fiscal_year or '?'})",
                    snippet=snippet,
                    source="historical_library",
                    citation=(
                        f"事务所 {r.firm_id} · 模板 {r.template_id} · "
                        f"脱敏入库 {r.uploaded_at:%Y-%m-%d}"
                    ),
                    score=min(0.95, score / 10.0),
                )
            )
        return hits
