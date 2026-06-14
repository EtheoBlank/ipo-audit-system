"""综合底稿 Excel 模板解析器。

读取符合 `docs/COMPREHENSIVE_WORKPAPER_TEMPLATE_SPEC.md` 的 .xlsx 模板，
输出结构化的 `TemplateSchema`。
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Optional, Union

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from app.services.comprehensive.schemas import TemplateField, TemplateSchema

logger = logging.getLogger(__name__)

# 单元格内占位符正则：{{ field_id }}
PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-z0-9_]+)\s*\}\}")

# _meta 表的字段定义表表头（与规范 4.2 一致）
FIELD_TABLE_HEADERS = {
    "field_id": "field_id",
    "label": "label",
    "type": "type",
    "source": "source",
    "required": "required",
    "hint": "hint",
    "options": "options",
}


class TemplateParseError(ValueError):
    """模板解析失败。"""


class TemplateParser:
    """综合底稿 Excel 模板解析器。

    用法:
        parser = TemplateParser()
        schema = parser.parse("templates/comprehensive/ar_comprehensive_v1.xlsx")
    """

    # ---------- 公共 API ----------

    def parse(self, source: Union[str, Path, bytes]) -> TemplateSchema:
        """从文件路径或字节流解析模板。"""
        wb = self._load_workbook(source)
        try:
            return self._parse_workbook(wb)
        finally:
            wb.close()

    # ---------- 内部 ----------

    @staticmethod
    def _load_workbook(source: Union[str, Path, bytes]):
        if isinstance(source, (str, Path)):
            return load_workbook(filename=str(source), data_only=False)
        from io import BytesIO

        return load_workbook(filename=BytesIO(source), data_only=False)

    def _parse_workbook(self, wb) -> TemplateSchema:
        meta_sheet = self._find_meta_sheet(wb)
        if meta_sheet is None:
            raise TemplateParseError(
                "未找到 _meta 工作表。请在模板中创建一个名为 '_meta' 的工作表（建议隐藏）。"
            )

        template_config = self._parse_template_config(meta_sheet)
        field_defs = self._parse_field_table(meta_sheet)

        # 用业务表中的占位符/命名区域核对 field_id 是否都登记
        business_placeholders = self._scan_business_placeholders(wb, exclude=meta_sheet.title)
        named_ranges = self._collect_named_ranges(wb, exclude=meta_sheet.title)

        # 用 _meta 表中登记的字段为权威；扫描结果只用来补充 cell_ref
        fields: list[TemplateField] = []
        seen: set[str] = set()
        for raw in field_defs:
            fid = raw["field_id"]
            if fid in seen:
                raise TemplateParseError(f"_meta 字段表中存在重复 field_id: {fid}")
            seen.add(fid)

            loc = self._locate_field(fid, business_placeholders, named_ranges)
            if loc is None:
                # 字段在 _meta 中登记，但模板中没有占位符 → 仍允许（视为"虚拟字段"）
                logger.warning(
                    "字段 '%s' 在 _meta 中登记但模板中未找到占位符/命名区域，"
                    "将作为虚拟字段保留（cell_ref 未知）。",
                    fid,
                )
                sheet_name = wb.sheetnames[0] if wb.sheetnames else "_meta"
                cell_ref = f"{sheet_name}!A1"
                row, col = 1, 1
                sheet = sheet_name
                name_range = None
            else:
                cell_ref, row, col, sheet, name_range = loc

            fields.append(
                TemplateField(
                    field_id=fid,
                    label=raw["label"],
                    type=raw.get("type", "text"),
                    source=raw["source"],
                    required=bool(raw.get("required", False)),
                    hint=raw.get("hint"),
                    options=raw.get("options"),
                    cell_ref=cell_ref,
                    name_range=name_range,
                    sheet=sheet,
                    row=row,
                    column=col,
                )
            )

        return TemplateSchema(
            template_id=template_config.get("template_id", "unknown"),
            template_name=template_config.get("template_name", "未命名模板"),
            version=template_config.get("version", "0.0.0"),
            firm_id=template_config.get("firm_id", "default"),
            industry=template_config.get("industry"),
            audit_period=template_config.get("audit_period"),
            required_workpapers=self._split_csv(template_config.get("required_workpapers")),
            manual_ref=template_config.get("manual_ref"),
            fields=fields,
            sheets=[s for s in wb.sheetnames if s != meta_sheet.title],
        )

    # ----- _meta 表解析 -----

    @staticmethod
    def _find_meta_sheet(wb) -> Optional[Any]:
        """查找 _meta 工作表。"""
        for name in wb.sheetnames:
            if name.strip().lower() == "_meta":
                return wb[name]
        return None

    @staticmethod
    def _parse_template_config(meta_sheet) -> dict[str, str]:
        """读取 A1:B? 区域的键值对配置。"""
        config: dict[str, str] = {}
        # 通常 config 在前 10 行；超过 10 行认为是字段表区域
        for row in meta_sheet.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
            key, value = row[0], row[1]
            if key is None or not isinstance(key, str):
                continue
            key = key.strip()
            if not key:
                continue
            # 命中字段表表头行就停
            if key.lower() in FIELD_TABLE_HEADERS:
                break
            config[key] = "" if value is None else str(value).strip()

        required = ("template_id", "template_name", "version", "firm_id")
        missing = [k for k in required if k not in config]
        if missing:
            raise TemplateParseError(f"_meta 配置缺少必填项: {missing}。需要: {required}")
        return config

    @staticmethod
    def _parse_field_table(meta_sheet) -> list[dict[str, Any]]:
        """读取字段定义表（A12 起，列: field_id, label, type, source, required, hint, options）。"""
        # 找到表头行
        header_row_idx: Optional[int] = None
        for i, row in enumerate(
            meta_sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1
        ):
            first = row[0]
            if isinstance(first, str) and first.strip().lower() == "field_id":
                header_row_idx = i
                break
        if header_row_idx is None:
            raise TemplateParseError(
                "_meta 中未找到字段定义表（缺少 'field_id' 表头）。"
                "请在 _meta 的某个区域按规范 4.2 的列顺序建立字段表。"
            )

        # 读表头映射
        header_cells = next(
            meta_sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)
        )
        header_map: dict[str, int] = {}
        for idx, h in enumerate(header_cells):
            if isinstance(h, str):
                key = h.strip().lower()
                if key in FIELD_TABLE_HEADERS:
                    header_map[key] = idx

        if "field_id" not in header_map or "source" not in header_map:
            raise TemplateParseError(
                f"字段定义表表头必须包含 field_id 和 source，实际: {list(header_map.keys())}"
            )

        # 读字段行
        rows: list[dict[str, Any]] = []
        for row in meta_sheet.iter_rows(
            min_row=header_row_idx + 1, max_row=meta_sheet.max_row, values_only=True
        ):
            if not row or row[header_map["field_id"]] is None:
                continue
            field_id = str(row[header_map["field_id"]]).strip()
            if not field_id:
                continue
            entry: dict[str, Any] = {"field_id": field_id}
            for key, idx in header_map.items():
                if key == "field_id":
                    continue
                v = row[idx] if idx < len(row) else None
                if v is None or (isinstance(v, str) and not v.strip()):
                    entry[key] = None if key in ("label", "type", "hint", "options") else False
                else:
                    entry[key] = v
            # required 兼容多种真值
            if "required" in entry:
                entry["required"] = str(entry["required"]).strip().lower() in (
                    "true",
                    "1",
                    "yes",
                    "y",
                    "是",
                    "√",
                )
            # type 缺省
            if not entry.get("type"):
                entry["type"] = "text"
            rows.append(entry)
        return rows

    # ----- 业务表占位符/命名区域扫描 -----

    def _scan_business_placeholders(self, wb, exclude: str) -> dict[str, tuple[str, int, int, str]]:
        """扫描所有业务表中 `{{field_id}}` 占位符的位置。

        Returns: {field_id: (cell_ref, row, column, sheet_name)}
        """
        found: dict[str, tuple[str, int, int, str]] = {}
        for name in wb.sheetnames:
            if name == exclude:
                continue
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    matches = PLACEHOLDER_RE.findall(val)
                    for fid in matches:
                        if fid in found:
                            logger.debug(
                                "占位符 '%s' 出现在多个位置，使用首次: %s 与 %s",
                                fid,
                                found[fid][0],
                                cell.coordinate,
                            )
                            continue
                        found[fid] = (f"{name}!{cell.coordinate}", cell.row, cell.column, name)
        return found

    @staticmethod
    def _collect_named_ranges(wb, exclude: str) -> dict[str, tuple[str, int, int, str]]:
        """收集命名区域到单元格坐标的映射。

        Returns: {name: (cell_ref, row, column, sheet_name)}
        """
        result: dict[str, tuple[str, int, int, str]] = {}

        # openpyxl 新版: wb.defined_names 是 Dict[str, DefinedName]
        defined = getattr(wb, "defined_names", None)
        if defined is None:
            return result

        def _iter_dns(obj):
            # 兼容不同版本
            if hasattr(obj, "items"):
                for name, dn in obj.items():
                    yield name, dn
            elif hasattr(obj, "definedName"):
                for dn in obj.definedName:
                    yield dn.name, dn

        for name, dn in _iter_dns(defined):
            try:
                destinations = list(dn.destinations)  # (sheet, cell_range)
            except Exception:  # noqa: BLE001
                continue
            for sheet_name, cell_range in destinations:
                if sheet_name == exclude:
                    continue
                # 只取区域左上角
                left_top = cell_range.split(":")[0]
                # 解析 row/col
                col_letter, row_num = coordinate_from_string(left_top)
                col_idx = column_index_from_string(col_letter)
                result[str(name)] = (f"{sheet_name}!{left_top}", row_num, col_idx, sheet_name)
        return result

    @staticmethod
    def _locate_field(
        field_id: str,
        placeholders: dict[str, tuple[str, int, int, str]],
        named_ranges: dict[str, tuple[str, int, int, str]],
    ) -> Optional[tuple[str, int, int, str, Optional[str]]]:
        """定位字段在模板中的位置：命名区域优先，否则占位符。"""
        if field_id in named_ranges:
            cell_ref, row, col, sheet = named_ranges[field_id]
            return cell_ref, row, col, sheet, field_id
        if field_id in placeholders:
            cell_ref, row, col, sheet = placeholders[field_id]
            return cell_ref, row, col, sheet, None
        return None

    @staticmethod
    def _split_csv(v: Optional[str]) -> list[str]:
        if not v:
            return []
        return [s.strip() for s in v.split(",") if s.strip()]
