"""Workbook generation service for audit workpapers."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
from typing import Iterable, Optional, Sequence

from app.core.config import settings

# 长期资产判定 — 与 account_audit 服务共用同一份默认前缀清单
from app.models.db.account_audit import DEFAULT_LONG_TERM_ASSET_PREFIXES


def _csv_prefixes(raw: Optional[str]) -> list[str]:
    if not raw:
        return []
    return [x.strip() for x in raw.split(",") if x.strip()]


def is_long_term_asset_account(
    account_code: str,
    extra_includes: Optional[Sequence[str]] = None,
    extra_excludes: Optional[Sequence[str]] = None,
) -> bool:
    """快速判定: 该科目是否长期资产 (用户特别要求, 发生额需逐笔审定).

    项目级覆盖需要由调用方传 ``extra_includes`` / ``extra_excludes``
    (来自 ``app.services.account_audit.get_effective_prefixes``).
    无传入时使用全局默认 + ``settings.LONG_TERM_ASSET_EXTRA_INCLUDES`` /
    ``LONG_TERM_ASSET_EXTRA_EXCLUDES``.
    """
    if not account_code:
        return False
    base = set(DEFAULT_LONG_TERM_ASSET_PREFIXES)
    base |= set(extra_includes or _csv_prefixes(settings.LONG_TERM_ASSET_EXTRA_INCLUDES))
    base -= set(extra_excludes or _csv_prefixes(settings.LONG_TERM_ASSET_EXTRA_EXCLUDES))
    code = str(account_code).strip()
    for p in base:
        if code.startswith(p):
            return True
    return False


class WorkbookGenerator:
    """Generate audit workbooks in Excel format."""

    def __init__(self, project_id: int, company_name: str, fiscal_year: int):
        self.project_id = project_id
        self.company_name = company_name
        self.fiscal_year = fiscal_year
        self.output_dir = settings.OUTPUT_DIR / f"project_{project_id}"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _get_styles(self) -> dict:
        """Get standard Excel styles for audit workbooks."""
        return {
            "header_font": Font(name="微软雅黑", size=11, bold=True, color="FFFFFF"),
            "header_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "subheader_font": Font(name="微软雅黑", size=10, bold=True),
            "subheader_fill": PatternFill(
                start_color="D6DCE5", end_color="D6DCE5", fill_type="solid"
            ),
            "title_font": Font(name="微软雅黑", size=14, bold=True),
            "normal_font": Font(name="微软雅黑", size=10),
            "thin_border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
            "center_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "left_align": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "right_align": Alignment(horizontal="right", vertical="center"),
        }

    def _apply_header_style(self, ws, row_num: int, col_num: int):
        """Apply header style to a cell."""
        styles = self._get_styles()
        cell = ws.cell(row=row_num, column=col_num)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = styles["thin_border"]

    def _apply_data_style(self, ws, row_num: int, col_num: int, is_number: bool = False):
        """Apply data cell style."""
        styles = self._get_styles()
        cell = ws.cell(row=row_num, column=col_num)
        cell.font = styles["normal_font"]
        cell.alignment = styles["right_align"] if is_number else styles["left_align"]
        cell.border = styles["thin_border"]

    def generate_account_detail(self, account_balances: pd.DataFrame) -> Path:
        """Generate account detail workbook (科目明细表).

        Args:
            account_balances: DataFrame with account balance data

        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "科目明细表"

        styles = self._get_styles()

        # Title
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"{self.company_name} - 科目明细表 ({self.fiscal_year}年度)"
        title_cell.font = styles["title_font"]
        title_cell.alignment = styles["center_align"]
        ws.row_dimensions[1].height = 30

        # Headers
        headers = [
            "科目编码",
            "科目名称",
            "期初余额",
            "借方发生额",
            "贷方发生额",
            "期末余额",
            "余额方向",
            "备注",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header)
            self._apply_header_style(ws, 2, col)
        ws.row_dimensions[2].height = 25

        # Data rows
        for row_idx, (_, row) in enumerate(account_balances.iterrows(), 3):
            ws.cell(row=row_idx, column=1, value=row.get("account_code", ""))
            ws.cell(row=row_idx, column=2, value=row.get("account_name", ""))
            ws.cell(row=row_idx, column=3, value=row.get("beginning_balance", 0))
            ws.cell(row=row_idx, column=4, value=row.get("debit_amount", 0))
            ws.cell(row=row_idx, column=5, value=row.get("credit_amount", 0))
            ws.cell(row=row_idx, column=6, value=row.get("ending_balance", 0))
            ws.cell(row=row_idx, column=7, value=row.get("balance_direction", ""))
            ws.cell(row=row_idx, column=8, value="")

            for col in range(1, 9):
                is_number = col in [3, 4, 5, 6]
                self._apply_data_style(ws, row_idx, col, is_number)

        # Set column widths
        col_widths = [15, 25, 15, 15, 15, 15, 10, 20]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width

        # Add summary row
        last_row = len(account_balances) + 3
        ws.cell(row=last_row, column=1, value="合计")
        ws.cell(row=last_row, column=1).font = styles["subheader_font"]
        ws.cell(row=last_row, column=3, value=account_balances["beginning_balance"].sum())
        ws.cell(row=last_row, column=4, value=account_balances["debit_amount"].sum())
        ws.cell(row=last_row, column=5, value=account_balances["credit_amount"].sum())
        ws.cell(row=last_row, column=6, value=account_balances["ending_balance"].sum())

        for col in range(1, 9):
            cell = ws.cell(row=last_row, column=col)
            cell.fill = styles["subheader_fill"]
            cell.border = styles["thin_border"]

        # Add filter
        ws.auto_filter.ref = f"A2:H{last_row - 1}"

        output_path = self.output_dir / f"科目明细表_{self.fiscal_year}.xlsx"
        wb.save(output_path)
        return output_path

    def generate_income_statement(self, account_balances: pd.DataFrame) -> Path:
        """Generate income statement workbook (利润表).

        Args:
            account_balances: DataFrame with account balance data

        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "利润表"

        styles = self._get_styles()

        # Title
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = f"{self.company_name} - 利润表 ({self.fiscal_year}年度)"
        title_cell.font = styles["title_font"]
        title_cell.alignment = styles["center_align"]
        ws.row_dimensions[1].height = 30

        # Headers
        headers = ["项目", "行次", "本期金额", "上期金额", "备注"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header)
            self._apply_header_style(ws, 2, col)
        ws.row_dimensions[2].height = 25

        # Income statement structure
        income_items = [
            ("营业收入", "5001"),
            ("营业成本", "5002"),
            ("研发费用", "5003"),
            ("销售费用", "5004"),
            ("管理费用", "5005"),
            ("财务费用", "5006"),
            ("公允价值变动收益", "5007"),
            ("投资收益", "5008"),
            ("其他收益", "5009"),
            ("营业利润", "5000"),
            ("加：营业外收入", "6001"),
            ("减：营业外支出", "6002"),
            ("利润总额", "6000"),
            ("减：所得税费用", "7001"),
            ("净利润", "7000"),
        ]

        # Filter income/expense accounts
        account_balances[account_balances["account_code"].str.startswith(("5", "6", "7"), na=False)]

        for row_idx, (item_name, item_code) in enumerate(income_items, 3):
            ws.cell(row=row_idx, column=1, value=item_name)
            ws.cell(row=row_idx, column=2, value=item_code)
            ws.cell(row=row_idx, column=3, value=0)
            ws.cell(row=row_idx, column=4, value=0)
            ws.cell(row=row_idx, column=5, value="")

            for col in range(1, 6):
                self._apply_data_style(ws, row_idx, col, col in [3, 4])

        # Set column widths
        col_widths = [20, 10, 18, 18, 20]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width

        output_path = self.output_dir / f"利润表_{self.fiscal_year}.xlsx"
        wb.save(output_path)
        return output_path

    def generate_balance_sheet(self, account_balances: pd.DataFrame) -> Path:
        """Generate balance sheet workbook (资产负债表).

        Args:
            account_balances: DataFrame with account balance data

        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "资产负债表"

        styles = self._get_styles()

        # Title
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = f"{self.company_name} - 资产负债表 ({self.fiscal_year}年度)"
        title_cell.font = styles["title_font"]
        title_cell.alignment = styles["center_align"]
        ws.row_dimensions[1].height = 30

        # Headers
        headers = ["项目", "行次", "期末余额", "年初余额", "备注"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header)
            self._apply_header_style(ws, 2, col)
        ws.row_dimensions[2].height = 25

        # Balance sheet structure

        output_path = self.output_dir / f"资产负债表_{self.fiscal_year}.xlsx"
        wb.save(output_path)
        return output_path

    def generate_cash_flow(self, chronological_accounts: pd.DataFrame) -> Path:
        """Generate cash flow statement workbook (现金流量表).

        Args:
            chronological_accounts: DataFrame with chronological account data

        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "现金流量表"

        styles = self._get_styles()

        # Title
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = f"{self.company_name} - 现金流量表 ({self.fiscal_year}年度)"
        title_cell.font = styles["title_font"]
        title_cell.alignment = styles["center_align"]
        ws.row_dimensions[1].height = 30

        # Headers
        headers = ["项目", "行次", "本期金额", "上期金额", "备注"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header)
            self._apply_header_style(ws, 2, col)
        ws.row_dimensions[2].height = 25

        output_path = self.output_dir / f"现金流量表_{self.fiscal_year}.xlsx"
        wb.save(output_path)
        return output_path

    def generate_trial_balance(
        self,
        account_balances: pd.DataFrame,
        consolidation_balances: pd.DataFrame = None,
    ) -> Path:
        """Generate trial balance workbook (试算平衡表) with consolidation columns.

        Args:
            account_balances: DataFrame with account balance data (单体报表)
            consolidation_balances: Optional DataFrame with consolidation balance data (合并报表)

        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "试算平衡表"

        styles = self._get_styles()

        # Title
        ws.merge_cells("A1:L1")
        title_cell = ws["A1"]
        title_cell.value = f"{self.company_name} - 试算平衡表 ({self.fiscal_year}年度)"
        title_cell.font = styles["title_font"]
        title_cell.alignment = styles["center_align"]
        ws.row_dimensions[1].height = 30

        # Headers - including consolidation columns if provided
        if consolidation_balances is not None:
            headers = [
                "科目编码",
                "科目名称",
                "期初借方(单)",
                "期初贷方(单)",
                "本期借方(单)",
                "本期贷方(单)",
                "期末借方(单)",
                "期末贷方(单)",
                "期末借方(合)",
                "期末贷方(合)",
                "内部抵销(资产)",
                "内部抵销(负债)",
            ]
        else:
            headers = [
                "科目编码",
                "科目名称",
                "期初借方",
                "期初贷方",
                "本期借方",
                "本期贷方",
                "期末借方",
                "期末贷方",
            ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header)
            self._apply_header_style(ws, 2, col)
        ws.row_dimensions[2].height = 30

        # Calculate trial balance for standalone
        debit_total_begin = account_balances[account_balances["balance_direction"] == "借"][
            "beginning_balance"
        ].sum()
        credit_total_begin = account_balances[account_balances["balance_direction"] == "贷"][
            "beginning_balance"
        ].sum()
        debit_total_current = account_balances["debit_amount"].sum()
        credit_total_current = account_balances["credit_amount"].sum()
        debit_total_end = account_balances[account_balances["balance_direction"] == "借"][
            "ending_balance"
        ].sum()
        credit_total_end = account_balances[account_balances["balance_direction"] == "贷"][
            "ending_balance"
        ].sum()

        # Calculate consolidation totals if provided
        cons_debit_end = 0
        cons_credit_end = 0
        if consolidation_balances is not None:
            cons_debit = consolidation_balances[consolidation_balances["balance_direction"] == "借"]
            cons_credit = consolidation_balances[
                consolidation_balances["balance_direction"] == "贷"
            ]
            cons_debit_end = cons_debit["ending_balance"].sum()
            cons_credit_end = cons_credit["ending_balance"].sum()

        # Summary section
        summary_row = 4
        ws.cell(row=summary_row, column=1, value="合计")
        ws.cell(row=summary_row, column=1).font = styles["subheader_font"]
        ws.cell(row=summary_row, column=3, value=debit_total_begin)
        ws.cell(row=summary_row, column=4, value=credit_total_begin)
        ws.cell(row=summary_row, column=5, value=debit_total_current)
        ws.cell(row=summary_row, column=6, value=credit_total_current)
        ws.cell(row=summary_row, column=7, value=debit_total_end)
        ws.cell(row=summary_row, column=8, value=credit_total_end)

        if consolidation_balances is not None:
            ws.cell(row=summary_row, column=9, value=cons_debit_end)
            ws.cell(row=summary_row, column=10, value=cons_credit_end)
            ws.cell(
                row=summary_row, column=11, value=debit_total_end - cons_debit_end
            )  # 内部抵销资产
            ws.cell(
                row=summary_row, column=12, value=credit_total_end - cons_credit_end
            )  # 内部抵销负债

        # Apply summary style
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=summary_row, column=col)
            cell.fill = styles["subheader_fill"]
            cell.border = styles["thin_border"]

        # Balance check
        check_row = 6
        is_balanced = abs(debit_total_end - credit_total_end) < 0.01
        ws.cell(row=check_row, column=1, value="单体平衡状态")
        ws.cell(row=check_row, column=2, value="✅平衡" if is_balanced else "❌ 不平衡")

        if consolidation_balances is not None:
            cons_is_balanced = abs(cons_debit_end - cons_credit_end) < 0.01
            ws.cell(row=check_row + 1, column=1, value="合并平衡状态")
            ws.cell(
                row=check_row + 1, column=2, value="✅ 平衡" if cons_is_balanced else "❌ 不平衡"
            )

            ws.cell(row=check_row + 2, column=1, value="内部交易抵销金额")
            ws.cell(row=check_row + 2, column=3, value=debit_total_end - cons_debit_end)
            ws.cell(row=check_row + 2, column=4, value=credit_total_end - cons_credit_end)

        # Set column widths
        col_widths = [15, 25, 12, 12, 12, 12, 12, 12]
        if consolidation_balances is not None:
            col_widths.extend([12, 12, 12, 12])
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[
                chr(64 + idx) if idx <= 26 else "A" + chr(64 + idx - 26)
            ].width = width

        # Freeze panes
        ws.freeze_panes = "A3"

        output_path = self.output_dir / f"试算平衡表_{self.fiscal_year}.xlsx"
        wb.save(output_path)
        return output_path

    # ============================================================
    #  审计说明 (调用知识库 + 法规库 + AI)
    # ============================================================

    def write_audit_notes_sheet(
        self,
        workbook_path: Path,
        notes: list[dict],
        sheet_name: str = "审计说明",
    ) -> Path:
        """在已有底稿 Excel 末尾追加一个"审计说明" sheet。

        Args:
            workbook_path: 已生成的底稿文件 (会被覆盖)
            notes: ``[{"account_code","account_name","note","references_kb","references_regulations"}]``
            sheet_name: 新 sheet 名

        Returns:
            写入后的文件路径
        """
        from openpyxl import load_workbook

        wb = load_workbook(workbook_path)
        # 同名 sheet 重写
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        styles = self._get_styles()

        # 标题
        ws.merge_cells("A1:E1")
        title = ws["A1"]
        title.value = f"{self.company_name} - 审计说明 ({self.fiscal_year}年度)"
        title.font = styles["title_font"]
        title.alignment = styles["center_align"]
        ws.row_dimensions[1].height = 28

        headers = ["科目编码", "科目名称", "审计说明", "知识库引用", "法规引用"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=h)
            self._apply_header_style(ws, 2, col)
        ws.row_dimensions[2].height = 24

        for r_idx, n in enumerate(notes, start=3):
            ws.cell(row=r_idx, column=1, value=n.get("account_code", ""))
            ws.cell(row=r_idx, column=2, value=n.get("account_name", ""))
            ws.cell(row=r_idx, column=3, value=n.get("note", ""))

            kb_refs = n.get("references_kb") or []
            kb_text = "\n".join(
                f"《{x.get('book_title', '')}》"
                + (f" / {x['chapter']}" if x.get("chapter") else "")
                + (f" / 第{x['page']}页" if x.get("page") else "")
                + f"  相似度 {x.get('score', 0):.2f}"
                for x in kb_refs
            )
            ws.cell(row=r_idx, column=4, value=kb_text)

            reg_refs = n.get("references_regulations") or []
            reg_text = "\n".join(
                f"《{x.get('title', '')}》"
                + (f" ({x['document_no']})" if x.get("document_no") else "")
                + (f"  {x['publish_date']}" if x.get("publish_date") else "")
                for x in reg_refs
            )
            ws.cell(row=r_idx, column=5, value=reg_text)

            for col in range(1, 6):
                cell = ws.cell(row=r_idx, column=col)
                cell.alignment = styles["left_align"]
                cell.border = styles["thin_border"]
                cell.font = styles["normal_font"]
            ws.row_dimensions[r_idx].height = max(40, min(220, len(str(n.get("note", ""))) // 3))

        # 列宽
        for col_idx, w in enumerate([14, 24, 70, 36, 36], 1):
            ws.column_dimensions[chr(64 + col_idx)].width = w
        ws.freeze_panes = "A3"

        wb.save(workbook_path)
        return workbook_path

    # ============================================================
    #  长期资产发生额审定 (用户特别要求 — Pack A)
    # ============================================================

    def generate_long_term_asset_workbook(
        self,
        account_code: str,
        account_name: str,
        period_end: str,
        balance_direction: str,
        beginning_book: float,
        beginning_audited: float,
        ending_book: float,
        ending_audited: float,
        debit_rows: Iterable[dict],
        credit_rows: Iterable[dict],
    ) -> Path:
        """生成长期资产专项底稿 (期初/借方/贷方/期末 都出审定数).

        长期资产科目要求: 不仅期初期末有审定数, 本期借方和贷方发生额也要逐笔审定。
        本方法生成的底稿包含 4 个 sheet:
          1. 封面 + 余额汇总(审定 vs 账面, 恒等式校验)
          2. 借方发生额明细 (审定逐笔)
          3. 贷方发生额明细 (审定逐笔)
          4. 调整分录汇总 (audited != book 的所有行)

        Args:
            account_code: 科目编码 (如 1601)
            account_name: 科目名称 (如 固定资产)
            period_end: 期末日期 YYYY-MM-DD
            balance_direction: 余额方向 借/贷
            beginning_book/audited: 期初账面/审定
            ending_book/audited: 期末账面/审定
            debit_rows: 借方发生额行, 每行 dict 含
                voucher_date / voucher_no / summary / counter_account /
                book_amount / audited_amount / adjustment_amount /
                adjustment_reason / working_paper_ref / status / audited_by_display
            credit_rows: 贷方发生额行 (字段同上)

        Returns:
            Path 生成的 Excel 文件
        """
        debit_list = list(debit_rows)
        credit_list = list(credit_rows)

        wb = Workbook()
        styles = self._get_styles()

        # === Sheet 1: 余额汇总 + 恒等式校验 ===
        ws_sum = wb.active
        ws_sum.title = "余额汇总"
        ws_sum.merge_cells("A1:G1")
        title = ws_sum["A1"]
        title.value = (
            f"{self.company_name} - {account_code} {account_name} 长期资产底稿 (期末 {period_end})"
        )
        title.font = styles["title_font"]
        title.alignment = styles["center_align"]
        ws_sum.row_dimensions[1].height = 30

        # 表头
        headers = ["项目", "账面金额", "审定金额", "审计调整", "调整方向", "占比", "备注"]
        for col, h in enumerate(headers, 1):
            ws_sum.cell(row=3, column=col, value=h)
            self._apply_header_style(ws_sum, 3, col)
        ws_sum.row_dimensions[3].height = 24

        debit_book_total = sum(float(r.get("book_amount", 0) or 0) for r in debit_list)
        debit_audited_total = sum(float(r.get("audited_amount", 0) or 0) for r in debit_list)
        credit_book_total = sum(float(r.get("book_amount", 0) or 0) for r in credit_list)
        credit_audited_total = sum(float(r.get("audited_amount", 0) or 0) for r in credit_list)

        is_debit_acc = balance_direction == "借"

        def _row(name: str, book: float, audited: float, ratio_base: float = 0.0) -> list:
            adj = audited - book
            direction = "增加" if adj > 0 else ("减少" if adj < 0 else "无")
            ratio = (audited / ratio_base * 100) if ratio_base else 0
            return [name, book, audited, adj, direction, f"{ratio:.2f}%" if ratio_base else "-", ""]

        rows = [
            _row("期初余额", beginning_book, beginning_audited),
            _row("本期借方发生额合计", debit_book_total, debit_audited_total),
            _row("本期贷方发生额合计", credit_book_total, credit_audited_total),
            _row("期末余额", ending_book, ending_audited),
        ]
        # 恒等式校验: 借方科目 期初 + 借 - 贷 - 期末 = 0; 贷方科目反过来
        if is_debit_acc:
            identity_book = beginning_book + debit_book_total - credit_book_total - ending_book
            identity_audited = (
                beginning_audited + debit_audited_total - credit_audited_total - ending_audited
            )
        else:
            identity_book = beginning_book - debit_book_total + credit_book_total - ending_book
            identity_audited = (
                beginning_audited - debit_audited_total + credit_audited_total - ending_audited
            )

        for r_idx, row in enumerate(rows, start=4):
            for col, val in enumerate(row, 1):
                ws_sum.cell(row=r_idx, column=col, value=val)
                self._apply_data_style(ws_sum, r_idx, col, is_number=col in (2, 3, 4))

        # 恒等式校验行 — 不平时标红
        check_row = 4 + len(rows) + 1
        ws_sum.cell(row=check_row, column=1, value="账面恒等式 (期初+借-贷-期末)")
        ws_sum.cell(row=check_row, column=2, value=round(identity_book, 2))
        ws_sum.cell(row=check_row + 1, column=1, value="审定恒等式 (审定值)")
        ws_sum.cell(row=check_row + 1, column=2, value=round(identity_audited, 2))
        balanced_book = abs(identity_book) < 0.01
        balanced_audited = abs(identity_audited) < 0.01
        ws_sum.cell(row=check_row, column=4, value="✅ 平衡" if balanced_book else "❌ 不平衡")
        ws_sum.cell(
            row=check_row + 1, column=4, value="✅ 平衡" if balanced_audited else "❌ 不平衡"
        )
        if not balanced_audited:
            for c in range(1, 8):
                cell = ws_sum.cell(row=check_row + 1, column=c)
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # 列宽
        for idx, w in enumerate([26, 18, 18, 18, 12, 12, 30], 1):
            ws_sum.column_dimensions[chr(64 + idx)].width = w

        # === Sheet 2 / 3: 借/贷方发生额明细 ===
        for sheet_name, source_rows in (
            ("借方发生额(审定)", debit_list),
            ("贷方发生额(审定)", credit_list),
        ):
            ws = wb.create_sheet(sheet_name)
            ws.merge_cells("A1:J1")
            t = ws["A1"]
            t.value = f"{account_code} {account_name} — {sheet_name} (期末 {period_end})"
            t.font = styles["title_font"]
            t.alignment = styles["center_align"]
            ws.row_dimensions[1].height = 28

            detail_headers = [
                "凭证日期",
                "凭证号",
                "摘要",
                "对方科目",
                "账面金额",
                "审定金额",
                "审计调整",
                "调整原因",
                "底稿索引",
                "审定状态",
            ]
            for col, h in enumerate(detail_headers, 1):
                ws.cell(row=2, column=col, value=h)
                self._apply_header_style(ws, 2, col)
            ws.row_dimensions[2].height = 26

            for r_idx, row in enumerate(source_rows, start=3):
                ws.cell(row=r_idx, column=1, value=row.get("voucher_date", ""))
                ws.cell(row=r_idx, column=2, value=row.get("voucher_no", ""))
                ws.cell(row=r_idx, column=3, value=row.get("summary", "") or "")
                ws.cell(row=r_idx, column=4, value=row.get("counter_account", "") or "")
                book = float(row.get("book_amount", 0) or 0)
                audited = float(row.get("audited_amount", 0) or 0)
                adj = audited - book
                ws.cell(row=r_idx, column=5, value=book)
                ws.cell(row=r_idx, column=6, value=audited)
                ws.cell(row=r_idx, column=7, value=adj)
                ws.cell(row=r_idx, column=8, value=row.get("adjustment_reason", "") or "")
                ws.cell(row=r_idx, column=9, value=row.get("working_paper_ref", "") or "")
                ws.cell(row=r_idx, column=10, value=row.get("status", "") or "")
                for col in range(1, 11):
                    self._apply_data_style(ws, r_idx, col, is_number=col in (5, 6, 7))
                # 调整非零的行高亮黄色
                if abs(adj) >= 0.01:
                    for col in range(1, 11):
                        cell = ws.cell(row=r_idx, column=col)
                        cell.fill = PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                        )

            # 合计行
            last = len(source_rows) + 3
            ws.cell(row=last, column=1, value="合计")
            ws.cell(
                row=last,
                column=5,
                value=sum(float(r.get("book_amount", 0) or 0) for r in source_rows),
            )
            ws.cell(
                row=last,
                column=6,
                value=sum(float(r.get("audited_amount", 0) or 0) for r in source_rows),
            )
            ws.cell(
                row=last,
                column=7,
                value=sum(
                    float(r.get("audited_amount", 0) or 0) - float(r.get("book_amount", 0) or 0)
                    for r in source_rows
                ),
            )
            for col in range(1, 11):
                cell = ws.cell(row=last, column=col)
                cell.fill = styles["subheader_fill"]
                cell.font = styles["subheader_font"]
                cell.border = styles["thin_border"]

            # 列宽
            for idx, w in enumerate([12, 14, 36, 14, 14, 14, 14, 28, 16, 12], 1):
                ws.column_dimensions[chr(64 + idx)].width = w
            ws.freeze_panes = "A3"
            if source_rows:
                ws.auto_filter.ref = f"A2:J{last - 1}"

        # === Sheet 4: 调整分录汇总 ===
        ws_adj = wb.create_sheet("调整分录汇总")
        ws_adj.merge_cells("A1:H1")
        t = ws_adj["A1"]
        t.value = f"{account_code} {account_name} — 调整分录 (期末 {period_end})"
        t.font = styles["title_font"]
        t.alignment = styles["center_align"]
        ws_adj.row_dimensions[1].height = 28

        adj_headers = ["凭证日期", "凭证号", "摘要", "方向", "账面", "审定", "调整", "调整原因"]
        for col, h in enumerate(adj_headers, 1):
            ws_adj.cell(row=2, column=col, value=h)
            self._apply_header_style(ws_adj, 2, col)

        adj_rows = []
        for direction_label, rows_iter in (("借", debit_list), ("贷", credit_list)):
            for r in rows_iter:
                book = float(r.get("book_amount", 0) or 0)
                audited = float(r.get("audited_amount", 0) or 0)
                if abs(audited - book) >= 0.01:
                    adj_rows.append(
                        {
                            "voucher_date": r.get("voucher_date", ""),
                            "voucher_no": r.get("voucher_no", ""),
                            "summary": r.get("summary", "") or "",
                            "direction": direction_label,
                            "book": book,
                            "audited": audited,
                            "adj": audited - book,
                            "reason": r.get("adjustment_reason", "") or "",
                        }
                    )

        for r_idx, r in enumerate(adj_rows, start=3):
            ws_adj.cell(row=r_idx, column=1, value=r["voucher_date"])
            ws_adj.cell(row=r_idx, column=2, value=r["voucher_no"])
            ws_adj.cell(row=r_idx, column=3, value=r["summary"])
            ws_adj.cell(row=r_idx, column=4, value=r["direction"])
            ws_adj.cell(row=r_idx, column=5, value=r["book"])
            ws_adj.cell(row=r_idx, column=6, value=r["audited"])
            ws_adj.cell(row=r_idx, column=7, value=r["adj"])
            ws_adj.cell(row=r_idx, column=8, value=r["reason"])
            for col in range(1, 9):
                self._apply_data_style(ws_adj, r_idx, col, is_number=col in (5, 6, 7))

        if not adj_rows:
            ws_adj.cell(row=3, column=1, value="(本期无调整分录)").font = styles["normal_font"]

        for idx, w in enumerate([12, 14, 36, 8, 14, 14, 14, 32], 1):
            ws_adj.column_dimensions[chr(64 + idx)].width = w
        ws_adj.freeze_panes = "A3"

        # 保存
        safe_name = "".join(c if c.isalnum() else "_" for c in account_name)[:40]
        output_path = self.output_dir / f"长期资产_{account_code}_{safe_name}_{period_end}.xlsx"
        # P0 第 2 轮修复 — wb.save 失败 (磁盘满/权限/路径) 不能让 API 500, 包成 IOError
        try:
            wb.save(output_path)
        except (OSError, IOError, PermissionError) as exc:
            import logging

            logging.getLogger(__name__).exception(
                "保存长期资产底稿失败 (account=%s, period=%s): %s",
                account_code,
                period_end,
                exc,
            )
            raise IOError(
                f"无法保存底稿到 {output_path}: {exc}. 请检查磁盘空间和目录权限."
            ) from exc
        return output_path
