"""WorkbookGenerator P0 修复测试 (2026-06-17).

覆盖 #7: generate_cash_flow 不再是空表, 而是按 经营/投资/筹资 三段分类.
        期末现金从 account_balances 取 1001/1002/1012 ending_balance 之和.
"""
from __future__ import annotations

import tempfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.core.config import settings
from app.services.workbook_generator import WorkbookGenerator


@pytest.fixture
def tmp_output_dir():
    """临时输出目录, 避免污染真实 outputs/."""
    with tempfile.TemporaryDirectory() as tmp:
        old = settings.OUTPUT_DIR
        settings.OUTPUT_DIR = Path(tmp)
        yield Path(tmp)
        settings.OUTPUT_DIR = old


def _ab(code: str, beginning: float = 0, ending: float = 0):
    return {
        "account_code": code,
        "beginning_balance": beginning,
        "ending_balance": ending,
        "debit_amount": 0,
        "credit_amount": 0,
    }


def _j(code: str, debit: float = 0, credit: float = 0, summary: str = ""):
    return {
        "account_code": code,
        "debit_amount": debit,
        "credit_amount": credit,
        "summary": summary,
    }


class TestCashFlowBasic:
    """基本三段 + 期初/期末."""

    def test_three_sections_present(self, tmp_output_dir):
        journals = pd.DataFrame([
            _j("1002", debit=100_000, summary="销售商品收到货款"),  # 经营 inflow
            _j("1002", credit=50_000, summary="购买商品支付货款"),  # 经营 outflow
            _j("1002", credit=30_000, summary="购建固定资产支付"),  # 投资 outflow
            _j("1002", debit=200_000, summary="取得银行借款"),  # 筹资 inflow
            _j("1002", credit=5_000, summary="偿还借款利息"),  # 筹资 outflow
        ])
        balances = pd.DataFrame([
            _ab("1001", beginning=5000, ending=8000),
            _ab("1002", beginning=100_000, ending=315_000),
        ])

        g = WorkbookGenerator(project_id=999, company_name="测试", fiscal_year=2024)
        out = g.generate_cash_flow(journals, balances)

        assert out.exists()
        wb = load_workbook(out)
        ws = wb.active

        # 找各关键行的"本期金额" (col 3)
        rows_by_label = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            label, _, amount, _, _ = row[:5]
            if label:
                rows_by_label[label] = amount

        # 经营: 流入 100k, 流出 50k, 净额 50k
        assert rows_by_label["经营活动现金流入小计"] == 100_000.0
        assert rows_by_label["经营活动现金流出小计"] == 50_000.0
        assert rows_by_label["经营活动产生的现金流量净额"] == 50_000.0
        # 投资: 流入 0, 流出 30k, 净额 -30k
        assert rows_by_label["投资活动现金流出小计"] == 30_000.0
        assert rows_by_label["投资活动产生的现金流量净额"] == -30_000.0
        # 筹资: 流入 200k, 流出 5k, 净额 195k
        assert rows_by_label["筹资活动现金流入小计"] == 200_000.0
        assert rows_by_label["筹资活动现金流出小计"] == 5_000.0
        assert rows_by_label["筹资活动产生的现金流量净额"] == 195_000.0
        # 净增加额 = 50 - 30 + 195 = 215k
        assert rows_by_label["四、现金及现金等价物净增加额"] == 215_000.0
        # 期初/期末 = 1001 + 1002 之和
        assert rows_by_label["加：期初现金及现金等价物余额"] == 105_000.0
        assert rows_by_label["五、期末现金及现金等价物余额"] == 323_000.0

    def test_without_account_balances(self, tmp_output_dir):
        """account_balances=None → 不写期初/期末, 不报错."""
        journals = pd.DataFrame([
            _j("1002", debit=50_000, summary="销售商品收到货款"),
        ])
        g = WorkbookGenerator(project_id=998, company_name="X", fiscal_year=2024)
        out = g.generate_cash_flow(journals, account_balances=None)

        wb = load_workbook(out)
        labels = [r[0] for r in wb.active.iter_rows(min_row=3, values_only=True) if r[0]]
        assert "经营活动现金流入小计" in labels
        assert "加：期初现金及现金等价物余额" not in labels  # 不写期初/期末
        assert "五、期末现金及现金等价物余额" not in labels

    def test_keyword_investing_classification(self, tmp_output_dir):
        """'购建固定资产' / '取得投资收' / '处置固定' → 投资.

        借贷方向: 1002 是现金科目, debit=现金流入, credit=现金流出.
        '购建固定资产' 是现金支出 (credit 100k), '处置固定资产收回' 是现金流入 (debit 50k).
        """
        journals = pd.DataFrame([
            _j("1002", credit=100_000, summary="购建固定资产支付现金"),  # 投资 outflow
            _j("1002", debit=50_000, summary="处置固定资产收回"),  # 投资 inflow
            _j("1002", credit=200_000, summary="投资支付现金"),  # 投资 outflow
            _j("1002", debit=80_000, summary="取得投资收益收到现金"),  # 投资 inflow
        ])
        g = WorkbookGenerator(project_id=997, company_name="X", fiscal_year=2024)
        out = g.generate_cash_flow(journals)

        wb = load_workbook(out)
        rows_by_label = {r[0]: r[2] for r in wb.active.iter_rows(min_row=3, values_only=True) if r[0]}
        # 投资: 流入 (处置+取得投资收) = 50k + 80k = 130k
        assert rows_by_label["投资活动现金流入小计"] == 130_000.0
        # 投资: 流出 (购建+投资支付) = 100k + 200k = 300k
        assert rows_by_label["投资活动现金流出小计"] == 300_000.0
        # 净额 = -170k
        assert rows_by_label["投资活动产生的现金流量净额"] == -170_000.0

    def test_keyword_financing_classification(self, tmp_output_dir):
        """'取得借款' / '偿还借款' / '分红' / '增资' → 筹资."""
        journals = pd.DataFrame([
            _j("1002", debit=500_000, summary="取得银行借款"),
            _j("1002", credit=100_000, summary="偿还借款本金"),
            _j("1002", credit=50_000, summary="分配股利支付现金"),
            _j("1002", debit=1_000_000, summary="吸收投资收到现金"),
        ])
        g = WorkbookGenerator(project_id=996, company_name="X", fiscal_year=2024)
        out = g.generate_cash_flow(journals)

        wb = load_workbook(out)
        rows_by_label = {r[0]: r[2] for r in wb.active.iter_rows(min_row=3, values_only=True) if r[0]}
        # 筹资: 流入 (取得借款+吸收投资) = 500k + 1000k = 1500k
        assert rows_by_label["筹资活动现金流入小计"] == 1_500_000.0
        # 流出 (偿还借款+分红) = 100k + 50k = 150k
        assert rows_by_label["筹资活动现金流出小计"] == 150_000.0

    def test_no_journals(self, tmp_output_dir):
        """空序时账 → 全 0, 不报错."""
        g = WorkbookGenerator(project_id=995, company_name="X", fiscal_year=2024)
        out = g.generate_cash_flow(pd.DataFrame(columns=["account_code", "debit_amount", "credit_amount", "summary"]))
        wb = load_workbook(out)
        rows_by_label = {r[0]: r[2] for r in wb.active.iter_rows(min_row=3, values_only=True) if r[0]}
        assert rows_by_label["经营活动现金流入小计"] == 0.0
        assert rows_by_label["经营活动产生的现金流量净额"] == 0.0
        assert rows_by_label["四、现金及现金等价物净增加额"] == 0.0

    def test_non_cash_accounts_ignored(self, tmp_output_dir):
        """非 1001/1002/1012 行 → 不计入现金流."""
        journals = pd.DataFrame([
            _j("5001", debit=100_000, summary="销售商品"),  # 收入科目, 不应计入
            _j("1002", debit=10_000, summary="其他经营活动"),  # 应计入
        ])
        g = WorkbookGenerator(project_id=994, company_name="X", fiscal_year=2024)
        out = g.generate_cash_flow(journals)
        wb = load_workbook(out)
        rows_by_label = {r[0]: r[2] for r in wb.active.iter_rows(min_row=3, values_only=True) if r[0]}
        assert rows_by_label["经营活动现金流入小计"] == 10_000.0

    def test_mismatch_warning_emitted(self, tmp_output_dir):
        """净增加额 ≠ 期末-期初 → 写警告."""
        # 故意造不平衡: 期末-期初 = 1000, 但 journal 算的净增加 = 5000
        journals = pd.DataFrame([
            _j("1002", debit=5000, summary="销售商品收到"),
        ])
        balances = pd.DataFrame([
            _ab("1002", beginning=1000, ending=2000),  # 差 1000
        ])
        g = WorkbookGenerator(project_id=993, company_name="X", fiscal_year=2024)
        out = g.generate_cash_flow(journals, balances)
        wb = load_workbook(out)
        notes = [r[4] for r in wb.active.iter_rows(min_row=3, values_only=True) if r[4]]
        # 至少有一条警告提到净增加额 ≠ 期末-期初
        assert any("净增加额" in str(n) and "期末-期初" in str(n) for n in notes)