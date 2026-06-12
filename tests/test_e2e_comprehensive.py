"""综合底稿端到端集成测试。

完整跑通：模板构建 → 解析 → 四路数据填充 → 问答补全 → Excel 导出。
"""
from __future__ import annotations

import asyncio
import io
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from app.services.comprehensive.builtin_rules import default_rule_book
from app.services.comprehensive.fill_engine import ComprehensiveFillEngine
from app.services.comprehensive.field_mapper import (
    FieldMapper,
    WorkpaperDataContext,
)
from app.services.comprehensive.qa_engine import QAEngine
from app.services.comprehensive.rule_engine import RuleEngine
from app.services.comprehensive.schemas import (
    FillReport,
    TemplateField,
    TemplateSchema,
)
from app.services.comprehensive.template_parser import TemplateParser
from app.services.comprehensive.web_search_engine import (
    SearchHit,
    WebSearchEngine,
)


# ============================================================
# 真实场景的"应收账款综合底稿"模板
# ============================================================

def _build_real_template() -> bytes:
    """构造一份接近实战的综合底稿模板。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "应收账款综合底稿"

    # 表头
    ws["A1"] = "公司：{{company_name}}"
    ws["A3"] = "审计期间：{{audit_period}}"
    ws["A5"] = "应收账款期末余额（元）：{{ar_balance}}"
    ws["A6"] = "应收账款周转天数：{{ar_turnover_days}}"
    ws["A7"] = "函证覆盖率：{{confirmation_rate}}"
    ws["A9"] = "风险等级：{{risk_level}}"
    ws["A11"] = "披露事项：{{disclosure_note}}"
    ws["A13"] = "管理层判断：{{mgmt_judgment}}"

    # 命名区域
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names["ar_balance"] = DefinedName(
        name="ar_balance", attr_text="'应收账款综合底稿'!$B$5"
    )
    wb.defined_names["confirmation_rate"] = DefinedName(
        name="confirmation_rate", attr_text="'应收账款综合底稿'!$B$7"
    )

    # _meta 表
    meta = wb.create_sheet("_meta")
    meta["A1"] = "template_id"; meta["B1"] = "ar_comprehensive_v1"
    meta["A2"] = "template_name"; meta["B2"] = "应收账款综合底稿（IPO）"
    meta["A3"] = "version"; meta["B3"] = "1.0.0"
    meta["A4"] = "firm_id"; meta["B4"] = "firm_demo"
    meta["A5"] = "industry"; meta["B5"] = "制造业"
    meta["A6"] = "audit_period"; meta["B6"] = "2024-01-01~2024-12-31"
    meta["A7"] = "required_workpapers"
    meta["B7"] = "应收账款明细表,函证汇总表,坏账准备计算表"
    meta["A8"] = "manual_ref"; meta["B8"] = "manual/ar_v3.md"

    meta["A12"] = "field_id"
    meta["B12"] = "label"
    meta["C12"] = "type"
    meta["D12"] = "source"
    meta["E12"] = "required"
    meta["F12"] = "hint"
    meta["G12"] = "options"

    field_rows = [
        ("company_name",        "公司全称",        "text",     "workpaper:project.company_name",       "true",  "",                 ""),
        ("audit_period",        "审计期间",        "text",     "workpaper:project.audit_period",       "true",  "YYYY-MM-DD~...",   ""),
        ("ar_balance",          "应收账款期末余额", "number",   "workpaper:ar_ledger.total_ending",     "true",  "单位:元",          ""),
        ("confirmation_rate",   "函证覆盖率",      "percent",  "workpaper:confirmation.coverage",      "true",  "0~1",              ""),
        ("ar_turnover_days",    "周转天数",        "number",   "calculated:365*ar_balance/revenue",    "true",  "自动",             ""),
        ("risk_level",          "风险等级",        "choice",   "rule:ar_risk_classify",                "true",  "",                 "低,中,高"),
        ("disclosure_note",     "披露事项",        "text_long","web_search:ar_disclosure",             "false", "CAS 22 披露",      ""),
        ("mgmt_judgment",       "管理层判断",      "text_long","human_qa",                             "true",  "≥200字",            ""),
    ]
    for i, row in enumerate(field_rows, start=13):
        for j, val in enumerate(row, start=1):
            meta.cell(row=i, column=j, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@dataclass
class _Project:
    company_name: str
    industry: str
    fiscal_year: int


def _build_context() -> WorkpaperDataContext:
    ab = pd.DataFrame([
        {"account_code": "1001", "account_name": "银行存款", "balance_direction": "借",
         "beginning_balance": 50000, "debit_amount": 10000, "credit_amount": 5000, "ending_balance": 55000},
        {"account_code": "1122", "account_name": "应收账款", "balance_direction": "借",
         "beginning_balance": 8000, "debit_amount": 4000, "credit_amount": 1000, "ending_balance": 11000},
        {"account_code": "2202", "account_name": "应付账款", "balance_direction": "贷",
         "beginning_balance": 5000, "debit_amount": 1000, "credit_amount": 2000, "ending_balance": 6000},
    ])

    @dataclass
    class _Conf:
        status: str
        confirmed_amount: float
        sent_amount: float = 0.0
        sample_balance: float = 0.0

    return WorkpaperDataContext(
        project=_Project("ACME 科技股份有限公司", "制造业", 2024),
        account_balances=ab,
        confirmation_cases=[
            _Conf("agreed", 6000, sent_amount=7000, sample_balance=11000),
            _Conf("agreed", 1000, sent_amount=2000, sample_balance=4000),
            _Conf("disputed", 500, sent_amount=1000, sample_balance=1000),
        ],
        extra={"revenue": 36500.0, "credit_sales": 30000.0},
    )


def _build_engine() -> ComprehensiveFillEngine:
    """构造一个能跑通 E2E 的引擎。"""

    async def fake_reg(q, k):
        if "披露" in q or "disclosure" in q.lower():
            return [SearchHit(
                title="CAS 22 应收账款披露要求",
                snippet="企业应当披露应收账账款的账面价值及减值准备、账龄分析、重大单项计提等。",
                source="", citation="财政部 · 财会〔2017〕7号",
                score=0.95,
            )]
        return []

    return ComprehensiveFillEngine(
        mapper=FieldMapper(),
        rule_engine=RuleEngine(default_rule_book()),
        web_engine=WebSearchEngine(regulation_search=fake_reg),
        qa_engine=QAEngine(),
    )


# ============================================================
# 端到端用例
# ============================================================

@pytest.mark.asyncio
async def test_e2e_full_pipeline():
    """端到端：从 Excel 模板 → 解析 → 填充 → 问答 → 导出 Excel。"""
    raw = _build_real_template()

    # 1) 解析
    schema = TemplateParser().parse(raw)
    assert schema.template_id == "ar_comprehensive_v1"
    assert len(schema.fields) == 8

    # 2) 跑四路填充
    engine = _build_engine()
    report = await engine.fill(schema, _build_context())
    by_id = {r.field_id: r for r in report.results}

    # workpaper
    assert by_id["company_name"].value == "ACME 科技股份有限公司"
    assert by_id["ar_balance"].value == 11000.0
    # confirmation coverage = sum(sent) / sum(sample) = 10000/16000 = 0.625
    assert abs(by_id["confirmation_rate"].value - 0.625) < 0.01

    # calculated：turnover_days = 365*ar_balance/revenue = 365*11000/36500 = 110
    assert by_id["ar_turnover_days"].value == pytest.approx(110.0)

    # rule：110 落在 [90, 120] → 中风险
    assert by_id["risk_level"].value == "中"
    assert by_id["risk_level"].source_used.startswith("rule:")

    # web_search：CAS 22 命中
    assert by_id["disclosure_note"].value is not None
    assert "披露" in by_id["disclosure_note"].value or "CAS" in by_id["disclosure_note"].value

    # 3) human_qa 必填 → 应进入 open_questions
    qa_topics = {q.topic for q in report.open_questions}
    assert "管理层判断" in qa_topics

    # 4) 模拟用户回答
    q = next(q for q in report.open_questions if q.topic == "管理层判断")
    final_report = await engine.apply_qa_answers(
        report, {q.question_id: "管理层已按账龄组合估计坏账，并经复核确认计提金额充分。"}
    )
    final_by_id = {r.field_id: r for r in final_report.results}
    assert final_by_id["mgmt_judgment"].value is not None
    assert "账龄" in final_by_id["mgmt_judgment"].value

    # 5) 导出 Excel（写回原模板）
    out_bytes = _export_with_log_sheet(raw, final_report)
    assert out_bytes

    wb = load_workbook(filename=io.BytesIO(out_bytes))
    ws = wb["应收账款综合底稿"]

    # 占位符被替换
    assert "ACME 科技股份有限公司" in ws["A1"].value
    assert "2024" in ws["A3"].value  # audit_period 含 2024
    # 命名区域被替换
    assert ws["B5"].value == 11000.0
    assert abs(ws["B7"].value - 0.625) < 0.01

    # _log 表存在并完整
    assert "_log" in wb.sheetnames
    log_rows = list(wb["_log"].values)
    assert len(log_rows) >= 9  # 1 header + 8 data
    field_ids = [r[0] for r in log_rows[1:]]
    for fid in ("company_name", "ar_balance", "ar_turnover_days",
                "risk_level", "disclosure_note", "mgmt_judgment"):
        assert fid in field_ids


@pytest.mark.asyncio
async def test_e2e_minimal_template_all_human_qa():
    """最小模板（全 human_qa 字段）：只有问答填得到。"""
    raw = _make_all_qa_template()
    schema = TemplateParser().parse(raw)
    assert all(f.source == "human_qa" for f in schema.fields)

    engine = _build_engine()
    report = await engine.fill(schema, _build_context())
    # 全部留作问答
    assert report.filled == 0
    assert report.pending == len(schema.fields)
    # 一次性问答把所有问题都给出
    answers = {q.question_id: f"测试回答 {i}"
               for i, q in enumerate(report.open_questions)}
    final = await engine.apply_qa_answers(report, answers)
    assert final.filled == len(schema.fields)
    assert final.pending == 0


@pytest.mark.asyncio
async def test_e2e_export_preserves_unrelated_cells():
    """导出后，未涉及占位符的单元格应原样保留。"""
    raw = _build_real_template()
    # 注入一个额外的非占位符单元格
    wb = load_workbook(filename=io.BytesIO(raw))
    wb["应收账款综合底稿"]["A20"] = "其他手工填写内容，不应被覆盖"
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    schema = TemplateParser().parse(raw)
    engine = _build_engine()
    report = await engine.fill(schema, _build_context())
    # 把所有问题回答掉
    answers = {q.question_id: "ok" for q in report.open_questions}
    final = await engine.apply_qa_answers(report, answers)

    out = _export_with_log_sheet(raw, final)
    wb2 = load_workbook(filename=io.BytesIO(out))
    assert wb2["应收账款综合底稿"]["A20"].value == "其他手工填写内容，不应被覆盖"


# ============================================================
# 辅助
# ============================================================

def _export_with_log_sheet(raw: bytes, report: FillReport) -> bytes:
    """前端 _export_to_excel 的等价实现（独立复刻以避免 streamlit 依赖）。"""
    wb = load_workbook(filename=io.BytesIO(raw))
    by_id = {r.field_id: r for r in report.results}

    # 重新解析以拿 cell 位置
    schema = TemplateParser().parse(raw)
    if "_log" in wb.sheetnames:
        del wb["_log"]
    log = wb.create_sheet("_log")
    log.append(["field_id", "sheet", "cell", "value", "source", "confidence", "citation"])

    for f in schema.fields:
        if f.field_id not in by_id or by_id[f.field_id].value is None:
            continue
        if f.sheet not in wb.sheetnames:
            continue
        ws = wb[f.sheet]
        cell = ws.cell(row=f.row, column=f.column)
        if isinstance(cell.value, str) and (
            cell.value.startswith("{{") or cell.value.endswith("}}")
        ):
            cell.value = by_id[f.field_id].value
        elif f.name_range:
            cell.value = by_id[f.field_id].value
        else:
            continue
        log.append([
            f.field_id, f.sheet, f.cell_ref,
            str(by_id[f.field_id].value),
            by_id[f.field_id].source_used,
            f"{by_id[f.field_id].confidence:.2f}",
            by_id[f.field_id].citation or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_all_qa_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Q"
    ws["A1"] = "{{a}}"
    ws["A2"] = "{{b}}"
    meta = wb.create_sheet("_meta")
    meta["A1"] = "template_id"; meta["B1"] = "qa"
    meta["A2"] = "template_name"; meta["B2"] = "Q"
    meta["A3"] = "version"; meta["B3"] = "1.0.0"
    meta["A4"] = "firm_id"; meta["B4"] = "f"
    meta["A12"] = "field_id"; meta["B12"] = "label"
    meta["C12"] = "type"; meta["D12"] = "source"
    meta["E12"] = "required"
    meta["A13"] = "a"; meta["B13"] = "A"; meta["C13"] = "text"
    meta["D13"] = "human_qa"; meta["E13"] = "true"
    meta["A14"] = "b"; meta["B14"] = "B"; meta["C14"] = "text"
    meta["D14"] = "human_qa"; meta["E14"] = "true"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
