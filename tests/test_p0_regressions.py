"""针对多 agent 审查发现的 P0 问题的回归测试。

每个测试对应一个或多个 P0 修复点，防止回归。
"""
from __future__ import annotations

import asyncio
import io
from dataclasses import dataclass
from typing import Any

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from app.services.comprehensive.builtin_rules import default_rule_book
from app.services.comprehensive.fill_engine import (
    ComprehensiveFillEngine,
    _safe_eval,
)
from app.services.comprehensive.field_mapper import (
    FieldMapper,
    WorkpaperDataContext,
)
from app.services.comprehensive.firm_template_service import (
    _anonymize_text,
    _replace_pat,
)
from app.services.comprehensive.qa_engine import QAEngine
from app.services.comprehensive.rule_engine import RuleEngine
from app.services.comprehensive.schemas import TemplateField
from app.services.comprehensive.template_parser import TemplateParser
from app.services.comprehensive.web_search_engine import (
    SearchHit,
    WebSearchEngine,
)


# ============================================================
# P0-1: 表达式引擎 AST 白名单（防注入）
# ============================================================

def test_safe_eval_rejects_def():
    with pytest.raises(ValueError):
        _safe_eval("def f(): return 1", {})


def test_safe_eval_rejects_lambda():
    with pytest.raises(ValueError):
        _safe_eval("(lambda: 1)()", {})


def test_safe_eval_rejects_walrus():
    with pytest.raises(ValueError):
        _safe_eval("(x := 5)", {})


def test_safe_eval_rejects_attribute_access():
    with pytest.raises(ValueError):
        _safe_eval("a.real", {"a": 1.0})


def test_safe_eval_rejects_call_to_non_whitelisted():
    with pytest.raises(ValueError):
        _safe_eval("open('x')", {})
    with pytest.raises(ValueError):
        _safe_eval("eval('1')", {})


def test_safe_eval_rejects_import():
    with pytest.raises(ValueError):
        _safe_eval("__import__('os')", {})


def test_safe_eval_rejects_list_comp():
    with pytest.raises(ValueError):
        _safe_eval("[x for x in [1,2]]", {})


def test_safe_eval_rejects_if_else():
    with pytest.raises(ValueError):
        _safe_eval("1 if True else 2", {})


def test_safe_eval_allows_arithmetic():
    assert _safe_eval("1 + 2 * 3", {}) == 7


def test_safe_eval_allows_comparison():
    assert _safe_eval("1 < 2", {}) is True
    assert _safe_eval("2 >= 2", {}) is True


def test_safe_eval_allows_bool():
    assert _safe_eval("True and False", {}) is False


def test_safe_eval_allows_whitelisted_function():
    assert _safe_eval("abs(-5)", {}) == 5
    assert _safe_eval("max(1, 2, 3)", {}) == 3


def test_safe_eval_rejects_long_string():
    with pytest.raises(ValueError):
        _safe_eval("a" * 3000, {})


def test_safe_eval_rejects_dunder_names():
    with pytest.raises(ValueError):
        _safe_eval("__class__", {})


# ============================================================
# P0-3: WebSearchEngine 三路并发
# ============================================================

@pytest.mark.asyncio
async def test_web_search_runs_sources_concurrently():
    """三路应并发执行（总耗时 < 单路耗时 × 3）。"""
    import time

    delay_per_source = 0.1

    async def slow_reg(q, k):
        await asyncio.sleep(delay_per_source)
        return [SearchHit(title=f"reg-{q}", snippet="x", source="", citation="")]
    async def slow_kb(q, k):
        await asyncio.sleep(delay_per_source)
        return [SearchHit(title=f"kb-{q}", snippet="x", source="", citation="")]
    async def slow_web(q, k):
        await asyncio.sleep(delay_per_source)
        return [SearchHit(title=f"web-{q}", snippet="x", source="", citation="")]

    engine = WebSearchEngine(
        regulation_search=slow_reg, kb_search=slow_kb, web_search=slow_web,
    )
    start = time.time()
    hits = await engine.search("test", top_k=5)
    elapsed = time.time() - start
    # 并发应接近单路耗时（而不是 3 倍）
    assert elapsed < delay_per_source * 2.5
    # 三路都有不同标题，去重后保留 3 条
    assert len(hits) == 3
    sources = {h.source for h in hits}
    assert sources == {"regulation", "knowledge_base", "web"}


@pytest.mark.asyncio
async def test_web_search_all_three_sources_fail_returns_empty():
    async def bad(q, k):
        raise RuntimeError("down")

    engine = WebSearchEngine(
        regulation_search=bad, kb_search=bad, web_search=bad,
    )
    hits = await engine.search("any", top_k=5)
    assert hits == []

    f = TemplateField(
        field_id="x", label="X", type="text", source="web_search:zzz",
        cell_ref="A1", sheet="s", row=1, column=1,
    )
    r = await engine.fill_field(f, {})
    assert r.value is None
    assert "未检索到" in (r.citation or "")


# ============================================================
# P0-5: 函证覆盖率口径
# ============================================================

def test_confirmation_coverage_uses_sent_over_sample():
    """coverage 分子应是 sent_amount，分母应是 sample_balance。"""
    from app.services.comprehensive.field_mapper import _resolve_confirmation, DataPath
    from app.services.comprehensive.field_mapper import WorkpaperDataContext

    @dataclass
    class C:
        status: str
        confirmed_amount: float = 0
        sent_amount: float = 0
        sample_balance: float = 0

    ctx = WorkpaperDataContext(confirmation_cases=[
        C("agreed", 6000, 7000, 11000),
        C("agreed", 1000, 2000, 4000),
    ])
    # coverage = (7000+2000)/(11000+4000) = 0.6
    assert _resolve_confirmation(DataPath("confirmation", ("coverage",)), ctx) == 0.6


def test_ar_turnover_prefers_credit_sales():
    """周转天数优先用 credit_sales。"""
    from app.services.comprehensive.field_mapper import _resolve_ledger, DataPath
    ab = pd.DataFrame([{
        "account_code": "1122", "account_name": "AR", "balance_direction": "借",
        "beginning_balance": 8000, "debit_amount": 4000, "credit_amount": 1000,
        "ending_balance": 11000,
    }])
    ctx = WorkpaperDataContext(
        account_balances=ab, extra={"credit_sales": 30000.0, "revenue": 50000.0}
    )
    r = _resolve_ledger(DataPath("ar_ledger", ("turnover_days",)), ctx, "1122")
    # 365 * 9500 / 30000 ≈ 115.58
    assert r == pytest.approx(115.58, rel=1e-3)


# ============================================================
# P0-6: 脱敏正则覆盖人名 / 英文 / 银行账号 / 信用代码
# ============================================================

def test_anonymize_replaces_chinese_person_name_with_title():
    counter: dict[str, int] = {}
    out = _anonymize_text("项目负责人：张三先生，审计师李四同志", counter)
    assert "张三" not in out
    assert "李四" not in out
    # 用 <PER_> 前缀
    assert "<PER_" in out


def test_anonymize_replaces_id_card():
    counter: dict[str, int] = {}
    id_card = "110101199003078811"
    out = _anonymize_text(f"身份证：{id_card}", counter)
    assert id_card not in out


def test_anonymize_replaces_bank_account():
    counter: dict[str, int] = {}
    out = _anonymize_text("银行卡：62226000123456789", counter)
    assert "62226000123456789" not in out


def test_anonymize_replaces_credit_code():
    counter: dict[str, int] = {}
    code = "91110000123456789X"
    out = _anonymize_text(f"统一社会信用代码：{code}", counter)
    assert code not in out


def test_anonymize_replaces_english_company():
    counter: dict[str, int] = {}
    out = _anonymize_text("Investor: Apple Inc. is a major shareholder", counter)
    assert "Apple Inc." not in out


def test_anonymize_replaces_audit_firm():
    counter: dict[str, int] = {}
    out = _anonymize_text("审计师事务所：普华永道中天会计师事务所", counter)
    assert "普华永道" not in out


def test_anonymize_replaces_mobile_phone():
    counter: dict[str, int] = {}
    out = _anonymize_text("联系方式：13800138000", counter)
    assert "13800138000" not in out


# ============================================================
# P0-2: /fill 不再是 silent failure
# ============================================================

@pytest.mark.asyncio
async def test_fill_handles_none_project_without_crash():
    """ctx.project = None 时不崩，且其他数据源仍能填。"""
    engine = ComprehensiveFillEngine(
        mapper=FieldMapper(),
        rule_engine=RuleEngine(default_rule_book()),
        web_engine=WebSearchEngine(),
        qa_engine=QAEngine(),
    )
    ctx = WorkpaperDataContext(
        account_balances=pd.DataFrame([{
            "account_code": "1122", "account_name": "AR", "balance_direction": "借",
            "beginning_balance": 0, "debit_amount": 0, "credit_amount": 0,
            "ending_balance": 0,
        }]),
        project=None,  # 关键
    )
    from app.services.comprehensive.schemas import TemplateSchema
    schema = TemplateSchema(
        template_id="t1", template_name="T", version="1.0.0", firm_id="f",
        fields=[
            TemplateField(
                field_id="x", label="X", type="number",
                source="workpaper:ar_ledger.total_ending",
                cell_ref="s!A1", sheet="s", row=1, column=1,
            ),
        ],
    )
    report = await engine.fill(schema, ctx)
    # 不崩，且 report 是合法对象
    assert report.template_id == "t1"
    assert report.total_fields == 1


# ============================================================
# P0-9: 模板 parser 重复 field_id / 命名区域指向 _meta
# ============================================================

def test_duplicate_field_id_rejected(tmp_path):
    """_meta 中两条 field_id 相同时抛 TemplateParseError。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "{{x}}"
    meta = wb.create_sheet("_meta")
    meta["A1"] = "template_id"; meta["B1"] = "t1"
    meta["A2"] = "template_name"; meta["B2"] = "T"
    meta["A3"] = "version"; meta["B3"] = "1.0.0"
    meta["A4"] = "firm_id"; meta["B4"] = "f"
    meta["A12"] = "field_id"; meta["B12"] = "label"
    meta["C12"] = "type"; meta["D12"] = "source"
    meta["E12"] = "required"
    meta["A13"] = "x"; meta["B13"] = "X"; meta["C13"] = "text"
    meta["D13"] = "human_qa"; meta["E13"] = "true"
    meta["A14"] = "x"; meta["B14"] = "X2"; meta["C14"] = "text"
    meta["D14"] = "human_qa"; meta["E14"] = "true"
    path = tmp_path / "dup.xlsx"
    wb.save(path)

    from app.services.comprehensive.template_parser import TemplateParseError
    with pytest.raises(TemplateParseError):
        TemplateParser().parse(path)


def test_named_range_pointing_to_meta_sheet_is_ignored():
    """命名区域指向 _meta sheet 时不影响业务解析。"""
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "{{x}}"
    meta = wb.create_sheet("_meta")
    meta["A1"] = "template_id"; meta["B1"] = "t1"
    meta["A2"] = "template_name"; meta["B2"] = "T"
    meta["A3"] = "version"; meta["B3"] = "1.0.0"
    meta["A4"] = "firm_id"; meta["B4"] = "f"
    meta["A12"] = "field_id"; meta["B12"] = "label"
    meta["C12"] = "type"; meta["D12"] = "source"
    meta["E12"] = "required"
    meta["A13"] = "x"; meta["B13"] = "X"; meta["C13"] = "text"
    meta["D13"] = "human_qa"; meta["E13"] = "true"
    # 命名区域指向 _meta
    wb.defined_names["stray"] = DefinedName(
        name="stray", attr_text="_meta!$A$1"
    )
    buf = io.BytesIO()
    wb.save(buf)
    schema = TemplateParser().parse(buf.getvalue())
    # 业务字段 1 个，stray 不计入
    assert len(schema.fields) == 1
    assert schema.fields[0].field_id == "x"


# ============================================================
# 同一主题 > 5 字段：合并为 1 问题
# ============================================================

@pytest.mark.asyncio
async def test_qa_merges_more_than_5_same_topic_into_one_question():
    """同一主题超过 5 个字段应合并为 1 个问题，field_ids 含全部。"""
    fields = [
        TemplateField(
            field_id=f"mgmt_{i}", label=f"判断 {i}", type="text_long",
            source="human_qa", required=True,
            cell_ref="A1", sheet="s", row=i+1, column=1,
        )
        for i in range(7)
    ]
    qa = QAEngine()
    qs = await qa.generate_questions(fields, filled_field_ids=set(), context={})
    # 7 个都属于"管理层判断"，合并为 1
    assert len(qs) == 1
    assert len(qs[0].field_ids) == 7
