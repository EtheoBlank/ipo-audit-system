"""Tests for comprehensive workpaper template parser."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from app.services.comprehensive.template_parser import TemplateParseError, TemplateParser


@pytest.fixture
def sample_template(tmp_path: Path) -> Path:
    """构造一份符合规范的最小可工作模板，写入 tmp_path 并返回路径。"""
    wb = Workbook()

    # ---- 业务表：应收账款综合底稿 ----
    ws = wb.active
    ws.title = "应收账款综合底稿"
    ws["A1"] = "公司：{{company_name}}"
    ws["A3"] = "审计期间：{{audit_period}}"
    ws["A5"] = "应收账款期末余额（元）：{{ar_balance}}"
    ws["A7"] = "应收账款周转天数：{{ar_turnover_days}}"
    ws["A9"] = "风险等级：{{risk_level}}"
    ws["A11"] = "披露事项：{{disclosure_note}}"
    ws["A13"] = "管理层判断：{{mgmt_judgment}}"

    # 命名区域
    wb.defined_names["ar_balance"] = DefinedName(
        name="ar_balance", attr_text="'应收账款综合底稿'!$B$5"
    )
    wb.defined_names["audit_period"] = DefinedName(
        name="audit_period", attr_text="'应收账款综合底稿'!$B$3"
    )

    # ---- _meta 表 ----
    meta = wb.create_sheet("_meta")
    meta["A1"] = "template_id"
    meta["B1"] = "ar_comprehensive_v1"
    meta["A2"] = "template_name"
    meta["B2"] = "应收账款综合底稿（示例）"
    meta["A3"] = "version"
    meta["B3"] = "1.0.0"
    meta["A4"] = "firm_id"
    meta["B4"] = "firm_demo"
    meta["A5"] = "industry"
    meta["B5"] = "制造业"
    meta["A6"] = "audit_period"
    meta["B6"] = "2024-01-01~2024-12-31"
    meta["A7"] = "required_workpapers"
    meta["B7"] = "应收账款明细表,函证汇总表,坏账准备计算表"
    meta["A8"] = "manual_ref"
    meta["B8"] = "manual/ar_v3.md"

    # 字段定义表
    meta["A12"] = "field_id"
    meta["B12"] = "label"
    meta["C12"] = "type"
    meta["D12"] = "source"
    meta["E12"] = "required"
    meta["F12"] = "hint"
    meta["G12"] = "options"

    field_rows = [
        ("company_name",        "公司全称",        "text",     "workpaper:project.company_name",       "true",  None,                                       None),
        ("audit_period",        "审计期间",        "text",     "workpaper:project.audit_period",       "true",  "格式: YYYY-MM-DD~YYYY-MM-DD",              None),
        ("ar_balance",          "应收账款期末余额", "number",   "workpaper:ar_ledger.total_ending",     "true",  "单位:元",                                  None),
        ("ar_turnover_days",    "应收账款周转天数", "number",   "calculated:365*ar_avg/revenue",        "true",  "自动计算",                                 None),
        ("risk_level",          "风险等级",        "choice",   "rule:ar_risk_classify",                "true",  None,                                       "低,中,高"),
        ("disclosure_note",     "披露事项",        "text_long","web_search:csrc_ar_disclosure",        "false", "引用 CAS 22 / 证监会披露要求",             None),
        ("mgmt_judgment",       "管理层判断",      "text_long","human_qa",                             "true",  "不少于 200 字，需描述坏账估计方法",         None),
    ]
    for i, row in enumerate(field_rows, start=13):
        for j, val in enumerate(row, start=1):
            meta.cell(row=i, column=j, value=val)

    path = tmp_path / "ar_comprehensive_v1.xlsx"
    wb.save(path)
    return path


# ---------- 用例 ----------

def test_parse_minimal_template(sample_template: Path):
    """能解析示例模板，得到 7 个字段 + 模板级配置。"""
    schema = TemplateParser().parse(sample_template)

    assert schema.template_id == "ar_comprehensive_v1"
    assert schema.template_name == "应收账款综合底稿（示例）"
    assert schema.version == "1.0.0"
    assert schema.firm_id == "firm_demo"
    assert schema.industry == "制造业"
    assert "应收账款明细表" in schema.required_workpapers
    assert "函证汇总表" in schema.required_workpapers
    assert schema.manual_ref == "manual/ar_v3.md"

    assert len(schema.fields) == 7
    fids = {f.field_id for f in schema.fields}
    assert fids == {
        "company_name", "audit_period", "ar_balance",
        "ar_turnover_days", "risk_level", "disclosure_note", "mgmt_judgment",
    }


def test_field_source_classification(sample_template: Path):
    """能按 source 前缀分桶。"""
    schema = TemplateParser().parse(sample_template)
    wp = schema.fields_by_source("workpaper")
    rule = schema.fields_by_source("rule")
    web = schema.fields_by_source("web_search")
    qa = schema.fields_by_source("human_qa")
    calc = schema.fields_by_source("calculated")

    assert {f.field_id for f in wp} == {"company_name", "audit_period", "ar_balance"}
    assert {f.field_id for f in rule} == {"risk_level"}
    assert {f.field_id for f in web} == {"disclosure_note"}
    assert {f.field_id for f in qa} == {"mgmt_judgment"}
    assert {f.field_id for f in calc} == {"ar_turnover_days"}


def test_named_range_takes_priority_over_placeholder(sample_template: Path):
    """ar_balance 既有占位符又有命名区域，cell_ref 应来自命名区域。"""
    schema = TemplateParser().parse(sample_template)
    f = schema.get_field("ar_balance")
    assert f is not None
    assert f.name_range == "ar_balance"
    assert f.cell_ref.endswith("!$B$5")
    assert f.sheet == "应收账款综合底稿"
    assert f.row == 5
    assert f.column == 2


def test_placeholder_only_field(sample_template: Path):
    """company_name 只有占位符、无命名区域，name_range 应为空。"""
    schema = TemplateParser().parse(sample_template)
    f = schema.get_field("company_name")
    assert f is not None
    assert f.name_range is None
    assert f.cell_ref.startswith("应收账款综合底稿!A1") or "A1" in f.cell_ref


def test_choice_field_options_parsed(sample_template: Path):
    """choice 类型的 options 字符串应被 split 成 list。"""
    schema = TemplateParser().parse(sample_template)
    f = schema.get_field("risk_level")
    assert f is not None
    assert f.type == "choice"
    assert f.options == ["低", "中", "高"]


def test_required_field_normalization(sample_template: Path):
    """required 接受 'true'/'是' 等中英文真值。"""
    schema = TemplateParser().parse(sample_template)
    for fid in ("company_name", "audit_period", "ar_balance", "mgmt_judgment"):
        assert schema.get_field(fid).required is True
    assert schema.get_field("disclosure_note").required is False


def test_missing_meta_sheet(tmp_path: Path):
    """没有 _meta 表 → 抛 TemplateParseError。"""
    wb = Workbook()
    wb.active["A1"] = "no meta"
    path = tmp_path / "no_meta.xlsx"
    wb.save(path)

    with pytest.raises(TemplateParseError):
        TemplateParser().parse(path)


def test_missing_template_config(tmp_path: Path):
    """_meta 缺 template_id → 抛错。"""
    wb = Workbook()
    meta = wb.active
    meta.title = "_meta"
    meta["A1"] = "version"
    meta["B1"] = "1.0.0"
    path = tmp_path / "bad_meta.xlsx"
    wb.save(path)

    with pytest.raises(TemplateParseError):
        TemplateParser().parse(path)


def test_invalid_source_rejected(tmp_path: Path):
    """字段的 source 不是合法前缀 → Pydantic 校验失败。"""
    from pydantic import ValidationError

    from app.services.comprehensive.schemas import TemplateField

    with pytest.raises(ValidationError):
        TemplateField(
            field_id="x",
            label="x",
            type="text",
            source="unknown:xxx",  # 非法
            cell_ref="A1",
            sheet="s",
            row=1,
            column=1,
        )
