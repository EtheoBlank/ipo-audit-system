"""End-to-end smoke test for the inventory module.

直接 import 服务模块（不启动 FastAPI），按以下链路跑一遍：
  1) 构造 movements DataFrame
  2) CountSheetBuilder.build
  3) InventoryAgingEngine.compute（含 NRV / 转回 / aging fallback）
  4) CountPhotoProcessor.match_to_sheets / completion_stats
  5) InventoryExporter.build → 写到 tmp 文件 → openpyxl 再读回来

验证：
  - 所有模块能正常 import 并实例化
  - 内部字段对得上（to_db_kwargs / row dict / ORM 字段）
  - 导出 xlsx 是合法 openpyxl 可读
"""

from __future__ import annotations

import io
import os
import sys
import tempfile
import traceback
from datetime import datetime, timedelta
from pathlib import Path

# 把项目根加进 sys.path，方便 import app.*
PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from openpyxl import load_workbook

from app.services.inventory import (
    CountPhotoProcessor,
    CountSheetBuilder,
    CountSheetStrategy,
    InventoryAgingEngine,
    InventoryExporter,
)
from app.services.inventory.count_sheet import CountSheetResult
from app.services.inventory.aging_engine import ImpairmentRow, ImpairmentResult
from app.services.inventory.photo_processor import ParsedCountRow


# 一个 stdout helper，不 print emoji（避免 Windows 控制台乱码）
def log(msg: str) -> None:
    try:
        print(f"[smoke] {msg}", flush=True)
    except UnicodeEncodeError:
        # Windows GBK stdout fallback
        safe = msg.encode("ascii", "replace").decode("ascii")
        print(f"[smoke] {safe}", flush=True)


def make_movements() -> list[dict]:
    """构造 3 个物料的本期收发存，混合新老批次。"""
    pe = datetime(2024, 12, 31)
    return [
        {  # 高单价，刚入的货
            "material_code": "M001",
            "material_name": "高精度螺丝",
            "category": "原材料",
            "spec": "M8x20",
            "unit": "个",
            "warehouse": "主仓",
            "batch_no": "B2024A",
            "inbound_date": pe - timedelta(days=30),
            "opening_qty": 0, "opening_amount": 0,
            "inbound_qty": 100, "inbound_amount": 2000,
            "outbound_qty": 0, "outbound_amount": 0,
            "ending_qty": 100, "ending_amount": 2000,
            "unit_cost": 20.0,
            "is_prior_year": False,
        },
        {  # 期初有 50 个老批次
            "material_code": "M002",
            "material_name": "钢板",
            "category": "原材料",
            "spec": "10mm",
            "unit": "吨",
            "warehouse": "二仓",
            "batch_no": "",
            "inbound_date": pe - timedelta(days=400),
            "opening_qty": 50, "opening_amount": 5000,
            "inbound_qty": 0, "inbound_amount": 0,
            "outbound_qty": 0, "outbound_amount": 0,
            "ending_qty": 50, "ending_amount": 5000,
            "unit_cost": 100.0,
            "is_prior_year": False,
        },
        {  # 9 个月前入的，应触发 181-365 分层
            "material_code": "M003",
            "material_name": "铜线",
            "category": "原材料",
            "spec": "2.5mm",
            "unit": "米",
            "warehouse": "主仓",
            "batch_no": "B2024C",
            "inbound_date": pe - timedelta(days=270),
            "opening_qty": 0, "opening_amount": 0,
            "inbound_qty": 200, "inbound_amount": 4000,
            "outbound_qty": 0, "outbound_amount": 0,
            "ending_qty": 200, "ending_amount": 4000,
            "unit_cost": 20.0,
            "is_prior_year": False,
        },
    ]


def make_sales() -> list:
    """构造期后销售清单，用于 NRV 测算。"""
    pe = datetime(2024, 12, 31)
    return [
        type("R", (), {
            "product_code": "M001",
            "revenue_confirm_date": pe + timedelta(days=15),
            "ship_date": None,
            "quantity": 10,
            "revenue_amount": 150,  # 低于账面 20
        })(),
        type("R", (), {
            "product_code": "M003",
            "revenue_confirm_date": pe + timedelta(days=20),
            "ship_date": None,
            "quantity": 50,
            "revenue_amount": 1100,  # 22 块，高于账面 20
        })(),
    ]


def main() -> int:
    log("=" * 60)
    log("Inventory module smoke test")
    log("=" * 60)

    # ----- 1) 构造 movements -----------------------------------------------
    movements = make_movements()
    log(f"[1] movements 构造完毕: {len(movements)} 行")

    # ----- 2) CountSheetBuilder.build --------------------------------------
    strategy = CountSheetStrategy(
        coverage_threshold=0.7,
        b_sample_ratio=0.2,
        c_sample_ratio=0.05,
        high_value_warehouses=["主仓"],
        must_include_categories=[],
        must_include_codes=[],
        min_unit_amount=0.0,
        random_seed=42,
    )
    cs_result: CountSheetResult = CountSheetBuilder.build(movements, strategy)
    assert isinstance(cs_result.rows, list), "rows 必须是 list"
    assert all(isinstance(r, dict) for r in cs_result.rows), "rows 每行必须是 dict"
    log(f"[2] CountSheetBuilder.build: rows={len(cs_result.rows)}, "
        f"coverage={cs_result.coverage_ratio:.2%}, "
        f"total_amount={cs_result.total_amount:.2f}")

    # 关键字段断言：to-be-ORM-mapped 字段都在
    required_row_keys = {
        "material_code", "material_name", "category", "warehouse", "batch_no", "unit",
        "book_qty", "book_unit_cost", "book_amount",
        "sample_tier", "sample_reason", "coverage_rank",
    }
    for r in cs_result.rows:
        missing = required_row_keys - set(r.keys())
        assert not missing, f"count-sheet row 缺字段: {missing} on {r}"
    log(f"    [2a] count-sheet row 字段齐全 ({len(required_row_keys)} 个)")

    # ----- 3) InventoryAgingEngine.compute --------------------------------
    engine = InventoryAgingEngine(industry="制造业", sell_cost_rate=0.05)
    sales = make_sales()
    imp_result: ImpairmentResult = engine.compute(
        movements,
        datetime(2024, 12, 31),
        sales_records=sales,
        prior_impairments={"M001": 200.0, "M002": 1000.0},
        manual_nrv={},
    )
    assert len(imp_result.rows) == 3, f"应有 3 行跌价结果，得到 {len(imp_result.rows)}"
    log(f"[3] AgingEngine.compute: rows={len(imp_result.rows)}, "
        f"summary keys={list(imp_result.summary.keys())}")

    # 验证 to_db_kwargs() 返回的所有 key 都是 ORM 字段
    expected_orm_fields = {
        "material_code", "material_name", "category", "period_end",
        "ending_qty", "book_unit_cost", "book_amount",
        "age_le_90", "age_91_180", "age_181_365", "age_366_730", "age_gt_730",
        "weighted_avg_age", "nrv_unit_price", "nrv_source", "nrv_amount",
        "estimated_sell_cost", "impairment_current", "impairment_opening",
        "impairment_reversal", "impairment_provision", "net_impairment_change",
        "method", "note",
    }
    row: ImpairmentRow = imp_result.rows[0]
    db_kwargs = row.to_db_kwargs()
    extra = set(db_kwargs) - expected_orm_fields
    missing = expected_orm_fields - set(db_kwargs)
    assert not extra, f"to_db_kwargs 多出 ORM 字段: {extra}"
    assert not missing, f"to_db_kwargs 缺少 ORM 字段: {missing}"
    log(f"    [3a] to_db_kwargs() ↔ ORM 字段一一对应 ({len(expected_orm_fields)} 个)")

    # ----- 4) CountPhotoProcessor -----------------------------------------
    # 把 count-sheet 假装为 ORM 行（用 type() 造一个轻量代理，仅需 match_to_sheets
    # 用到的字段）
    class _Sheet:
        def __init__(self, **kw):
            for k, v in kw.items():
                setattr(self, k, v)

    sheets = []
    for r in cs_result.rows:
        s = _Sheet(
            material_code=r["material_code"],
            material_name=r["material_name"],
            warehouse=r["warehouse"],
            batch_no=r["batch_no"],
            book_qty=r["book_qty"],
            book_unit_cost=r["book_unit_cost"],
            book_amount=r["book_amount"],
            counted_qty=None,
        )
        sheets.append(s)
    log(f"[4a] count sheets 准备好: {len(sheets)} 行")

    parsed = [
        ParsedCountRow(material_code="M001", material_name="高精度螺丝",
                       counted_qty=99, warehouse="主仓", batch_no="B2024A"),
        ParsedCountRow(material_code="M002", material_name="钢板",
                       counted_qty=50, warehouse="二仓"),
        ParsedCountRow(material_code="M003", material_name="铜线",
                       counted_qty=195, warehouse="主仓"),
        ParsedCountRow(material_code="MXXX", material_name="未知物料",
                       counted_qty=10),
    ]
    matched, unmatched = CountPhotoProcessor.match_to_sheets(parsed, sheets)
    assert len(matched) == 3, f"应匹配 3 行，匹配了 {len(matched)}"
    assert len(unmatched) == 1, f"应未匹配 1 行，未匹配 {len(unmatched)}"
    log(f"[4] match_to_sheets: matched={len(matched)}, unmatched={len(unmatched)}")

    # 回填 counted_qty
    for s, p in matched:
        s.counted_qty = p.counted_qty

    # completion_stats
    stats = CountPhotoProcessor.completion_stats(sheets)
    assert stats["overall"]["total_items"] == len(sheets)
    assert stats["overall"]["counted_items"] == 3
    log(f"[4b] completion_stats: items_rate={stats['overall']['items_rate']:.2%}, "
        f"diff_total={stats['difference_summary']['total_count']}")

    # ----- 5) InventoryExporter.build → 写 tmp → openpyxl 再读 ------------
    class _CountPlan:
        title = "测试公司 2024 年度存货监盘计划"
        industry = "制造业"
        period_end = "2024-12-31"
        count_date_start = "2024-12-31"
        count_date_end = "2025-01-02"
        objectives = "测试目标"
        scope = "测试范围"
        team = '[{"name":"审计经理","role":"现场负责人","contact":""}]'
        procedures = "测试程序"
        special_notes = "测试特殊事项"
        risks = "测试重大风险"

    blob = InventoryExporter.build(
        movements=movements,
        count_sheets=sheets,
        plan=_CountPlan(),
        completion=stats,
        impairments=imp_result.rows,
        summary=imp_result.summary,
    )
    assert isinstance(blob, bytes) and len(blob) > 1000, f"导出文件异常: {len(blob)} bytes"

    # 写到 tmp，再用 openpyxl 读回来
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tf.write(blob)
        tmp_path = tf.name
    log(f"[5] InventoryExporter.build: {len(blob)} bytes 写入 {tmp_path}")

    try:
        wb = load_workbook(tmp_path, read_only=True)
        sheet_names = wb.sheetnames
        log(f"    [5a] openpyxl 再读成功, sheets = {sheet_names}")
        expected_sheets = {"收发存明细", "盘点计划", "盘点用表", "已盘点情况",
                           "盘点率统计", "库龄分析", "跌价测试", "跌价汇总"}
        missing_sheets = expected_sheets - set(sheet_names)
        assert not missing_sheets, f"导出 xlsx 缺 sheet: {missing_sheets}"
        # 抽样读几行
        for name in ["盘点用表", "跌价测试"]:
            ws = wb[name]
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter)
            n_data = sum(1 for _ in rows_iter)
            log(f"    [5b] sheet '{name}': {len(header)} 列, {n_data} 行数据")
        wb.close()
    finally:
        os.unlink(tmp_path)

    log("=" * 60)
    log("ALL SMOKE STEPS PASSED")
    log("=" * 60)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except AssertionError as exc:
        log(f"ASSERT FAILED: {exc}")
        traceback.print_exc()
        sys.exit(1)
    except Exception as exc:
        log(f"UNEXPECTED ERROR: {exc}")
        traceback.print_exc()
        sys.exit(2)
