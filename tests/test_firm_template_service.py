"""Tests for multi-firm template + historical workpaper library."""
from __future__ import annotations

from io import BytesIO
from typing import Any

import pytest
import pytest_asyncio
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

from app.core.database import Base
from app.models.db_models import FirmTemplate, HistoricalWorkpaper
from app.services.comprehensive.firm_template_service import (
    FirmTemplateService,
    HistoricalLibraryService,
    _anonymize_excel,
    _anonymize_text,
)


# ---------- in-memory SQLite for tests ----------

@pytest_asyncio.fixture
async def session() -> AsyncSession:
    # StaticPool：所有 session 共享同一连接 → in-memory DB 可见；
    # 同时避免 aiosqlite 事件循环冲突。
    engine = create_async_engine(
        "sqlite+aiosqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    sm = async_sessionmaker(engine, expire_on_commit=False)
    async with sm() as s:
        yield s
    await engine.dispose()


def _make_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "{{x}}"
    meta = wb.create_sheet("_meta")
    meta["A1"] = "template_id"; meta["B1"] = "t1"
    meta["A2"] = "template_name"; meta["B2"] = "T"
    meta["A3"] = "version"; meta["B3"] = "1.0.0"
    meta["A4"] = "firm_id"; meta["B4"] = "fA"
    # 字段定义表（最小可工作）
    meta["A12"] = "field_id"
    meta["B12"] = "label"
    meta["C12"] = "type"
    meta["D12"] = "source"
    meta["E12"] = "required"
    meta["A13"] = "x"
    meta["B13"] = "X"
    meta["C13"] = "text"
    meta["D13"] = "human_qa"
    meta["E13"] = "false"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_workpaper_bytes() -> bytes:
    """含敏感企业名的历史底稿。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "审计客户：北京蓝色星际科技股份有限公司"
    ws["A2"] = "其他：北京蓝色星际科技股份有限公司与上海XX集团有关联交易"
    ws["A3"] = "披露：根据 CAS 22，AA 集团应披露"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- _anonymize_text ----------

def test_anonymize_replaces_company_names():
    text = "北京蓝色星际科技股份有限公司的关联方为上海华兴集团有限公司"
    counter: dict[str, int] = {}
    out = _anonymize_text(text, counter)
    assert "北京蓝色星际" not in out
    assert "上海华兴" not in out
    assert "<ENT_1>" in out
    assert "<ENT_2>" in out


def test_anonymize_dedups_same_company_to_same_token():
    text = "蓝色星际集团和蓝色星际集团是关联方"
    counter: dict[str, int] = {}
    out = _anonymize_text(text, counter)
    # 同一公司多次出现应映射到同一 token
    tokens = [t for t in out.split() if t.startswith("<ENT_")]
    assert len(set(tokens)) == 1


def test_anonymize_excel_returns_bytes_and_excerpt():
    raw = _make_workpaper_bytes()
    out_bytes, excerpt = _anonymize_excel(raw)
    assert isinstance(out_bytes, bytes)
    assert len(out_bytes) > 0
    # 摘要中不应包含原始公司名
    assert "北京蓝色星际" not in excerpt
    assert "上海华兴" not in excerpt
    # 但 CAS 22 这类术语保留
    assert "CAS 22" in excerpt


# ---------- FirmTemplateService ----------

@pytest.mark.asyncio
async def test_upload_template_creates_record(session: AsyncSession):
    svc = FirmTemplateService(session)
    t = await svc.upload(
        firm_id="firmA", template_id="t1", version="1.0.0",
        template_bytes=_make_template_bytes(),
        template_name="T1",
    )
    assert t.id is not None
    assert t.firm_id == "firmA"
    assert t.is_active is True
    # field_schema_json 已生成
    assert t.field_schema_json is not None
    assert "{{x}}" in t.field_schema_json or "x" in t.field_schema_json


@pytest.mark.asyncio
async def test_upload_same_version_overwrites(session: AsyncSession):
    svc = FirmTemplateService(session)
    t1 = await svc.upload(
        firm_id="firmA", template_id="t1", version="1.0.0",
        template_bytes=_make_template_bytes(), template_name="T1 v1",
    )
    t1_id = t1.id
    t2 = await svc.upload(
        firm_id="firmA", template_id="t1", version="1.0.0",
        template_bytes=_make_template_bytes(), template_name="T1 v1 updated",
    )
    assert t2.id == t1_id  # 覆盖而非新建


@pytest.mark.asyncio
async def test_list_for_firm_isolates_firms(session: AsyncSession):
    svc = FirmTemplateService(session)
    await svc.upload("firmA", "t1", "1.0.0", _make_template_bytes(), "T1")
    await svc.upload("firmA", "t2", "1.0.0", _make_template_bytes(), "T2")
    await svc.upload("firmB", "t1", "1.0.0", _make_template_bytes(), "Other T1")

    a = await svc.list_for_firm("firmA")
    b = await svc.list_for_firm("firmB")
    assert {t.template_id for t in a} == {"t1", "t2"}
    assert {t.template_id for t in b} == {"t1"}


@pytest.mark.asyncio
async def test_get_latest_returns_most_recent_version(session: AsyncSession):
    svc = FirmTemplateService(session)
    await svc.upload("firmA", "t1", "1.0.0", _make_template_bytes(), "old")
    await svc.upload("firmA", "t1", "2.0.0", _make_template_bytes(), "new")
    t = await svc.get_latest("firmA", "t1")
    assert t is not None
    assert t.version == "2.0.0"


@pytest.mark.asyncio
async def test_deactivate_soft_deletes(session: AsyncSession):
    svc = FirmTemplateService(session)
    await svc.upload("firmA", "t1", "1.0.0", _make_template_bytes(), "T1")
    assert await svc.deactivate("firmA", "t1", "1.0.0") is True
    assert await svc.get_latest("firmA", "t1") is None  # 找不到 active


@pytest.mark.asyncio
async def test_parse_to_schema_round_trip(session: AsyncSession):
    svc = FirmTemplateService(session)
    await svc.upload("firmA", "t1", "1.0.0", _make_template_bytes(), "T1")
    schema = await svc.parse_to_schema("firmA", "t1")
    assert schema is not None
    # firm_id 来自 _meta 表（模板内嵌）
    assert schema.firm_id == "fA"
    assert len(schema.fields) == 1
    assert schema.fields[0].field_id == "x"


# ---------- HistoricalLibraryService ----------

@pytest.mark.asyncio
async def test_ingest_anonymizes_and_stores(session: AsyncSession):
    svc = HistoricalLibraryService(session)
    rec = await svc.ingest(
        firm_id="firmA", template_id="t1",
        workpaper_bytes=_make_workpaper_bytes(),
        industry="制造业", fiscal_year=2023,
    )
    assert rec.id is not None
    assert rec.firm_id == "firmA"
    assert rec.source_project_hash is not None
    # 脱敏后字节中不应含原公司名
    text = rec.text_excerpt or ""
    assert "北京蓝色星际" not in text
    assert "上海华兴" not in text


@pytest.mark.asyncio
async def test_search_finds_keyword(session: AsyncSession):
    svc = HistoricalLibraryService(session)
    await svc.ingest("firmA", "t1", _make_workpaper_bytes())
    hits = await svc.search("firmA", "t1", "CAS 22")
    assert len(hits) >= 1
    assert hits[0].source == "historical_library"
    assert "firmA" in hits[0].citation


@pytest.mark.asyncio
async def test_search_isolated_by_firm(session: AsyncSession):
    svc = HistoricalLibraryService(session)
    await svc.ingest("firmA", "t1", _make_workpaper_bytes())
    await svc.ingest("firmB", "t1", _make_workpaper_bytes())

    a_hits = await svc.search("firmA", "t1", "CAS")
    b_hits = await svc.search("firmB", "t1", "CAS")
    assert all("firmA" in h.citation for h in a_hits)
    assert all("firmB" in h.citation for h in b_hits)


@pytest.mark.asyncio
async def test_search_returns_empty_for_no_match(session: AsyncSession):
    svc = HistoricalLibraryService(session)
    await svc.ingest("firmA", "t1", _make_workpaper_bytes())
    hits = await svc.search("firmA", "t1", "完全不存在的关键词xyz")
    assert hits == []


@pytest.mark.asyncio
async def test_search_ignores_short_keywords(session: AsyncSession):
    svc = HistoricalLibraryService(session)
    await svc.ingest("firmA", "t1", _make_workpaper_bytes())
    # 长度 < 2 的词被忽略
    hits = await svc.search("firmA", "t1", "a I")
    assert hits == []
