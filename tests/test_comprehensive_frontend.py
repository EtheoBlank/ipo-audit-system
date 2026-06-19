"""Tests for the comprehensive workpaper frontend (non-Streamlit parts).

完整跑 Streamlit 需浏览器集成；这里只测纯函数与导出逻辑。
"""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

from app.services.comprehensive.schemas import FillResult, TemplateField, TemplateSchema


# ---------- 内部辅助：构造一个能跑 _export_to_excel 的最小场景 ----------

def _build_template_with_placeholders() -> bytes:
    """构造一个含占位符和 _meta 表的最小模板。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "{{a}}"
    ws["B1"] = "{{b}}"
    ws["A2"] = "static"

    meta = wb.create_sheet("_meta")
    meta["A1"] = "template_id"; meta["B1"] = "t1"
    meta["A2"] = "template_name"; meta["B2"] = "T"
    meta["A3"] = "version"; meta["B3"] = "1.0.0"
    meta["A4"] = "firm_id"; meta["B4"] = "f"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_schema() -> TemplateSchema:
    return TemplateSchema(
        template_id="t1", template_name="T", version="1.0.0", firm_id="f",
        sheets=["S"],
        fields=[
            TemplateField(
                field_id="a", label="A", type="text", source="workpaper:x",
                cell_ref="S!A1", sheet="S", row=1, column=1,
            ),
            TemplateField(
                field_id="b", label="B", type="text", source="human_qa",
                cell_ref="S!B1", sheet="S", row=1, column=2,
            ),
        ],
    )


def _make_report() -> "FillReport":  # type: ignore[name-defined]
    from app.services.comprehensive.schemas import FillReport
    # 用 dict 而不是 FillResult 实例 —— 全量 pytest 跑下,
    # `app.services.comprehensive.schemas` 可能因 import 顺序被加载两次,
    # 导致 Pydantic v2 严格类身份校验把"同一个类的另一份"当成异类拒绝.
    # 用 dict 走 Pydantic 的常规校验通道,既等价又避免假阳性.
    return FillReport(
        template_id="t1", total_fields=2, filled=2, pending=0,
        results=[
            {"field_id": "a", "value": "alpha", "source_used": "workpaper:x",
             "confidence": 0.95, "citation": "…"},
            {"field_id": "b", "value": "beta", "source_used": "human_qa:1",
             "confidence": 1.0, "citation": "…"},
        ],
        open_questions=[],
    )


def test_export_writes_values_and_creates_log_sheet():
    """把 FillReport 写回原模板，并在 _log 表中记录来源。

    2026-06-19 P1 修复: 旧版 _export_to_excel 读 __comprehensive_template_bytes__
    (永远空), 改成读 __comprehensive_template_path__ (实际有 tempfile).
    """
    from frontend.pages_comprehensive import _export_to_excel  # type: ignore

    # 把模板字节写到 tempfile, path 放进 session_state
    import tempfile
    from pathlib import Path
    raw = _build_template_with_placeholders()
    tmp_dir = Path(tempfile.gettempdir())
    tmp_path = tmp_dir / "test_export_template.xlsx"
    tmp_path.write_bytes(raw)
    try:
        import streamlit as st  # noqa: F401
        class _SS(dict):
            def get(self, k, default=None):
                return super().get(k, default)
        st.session_state = _SS({
            "__comprehensive_template_path__": str(tmp_path),
        })

        out = _export_to_excel(_make_schema(), _make_report())
        assert out  # non-empty
    finally:
        if tmp_path.exists():
            tmp_path.unlink()

    wb = load_workbook(filename=io.BytesIO(out))
    ws = wb["S"]
    # A1 原来是 {{a}} → "alpha"
    assert ws["A1"].value == "alpha"
    # B1 原来是 {{b}} → "beta"
    assert ws["B1"].value == "beta"
    # _log 表存在并记录了填充轨迹
    assert "_log" in wb.sheetnames
    log_rows = list(wb["_log"].values)
    # 1 行表头 + 2 行数据
    assert len(log_rows) == 3
    headers = log_rows[0]
    assert "field_id" in headers
    assert "source" in headers
    # 两条数据包含 field_id
    field_ids_in_log = [r[0] for r in log_rows[1:]]
    assert "a" in field_ids_in_log
    assert "b" in field_ids_in_log


def test_export_returns_empty_when_no_template_bytes():
    from frontend.pages_comprehensive import _export_to_excel  # type: ignore
    import streamlit as st
    st.session_state = {}  # no bytes
    out = _export_to_excel(_make_schema(), _make_report())
    assert out == b""


# ---------- _truncate ----------

def test_truncate_short_string_unchanged():
    from frontend.pages_comprehensive import _truncate  # type: ignore
    assert _truncate("abc", 10) == "abc"


def test_truncate_long_string_gets_ellipsis():
    from frontend.pages_comprehensive import _truncate  # type: ignore
    s = "x" * 100
    out = _truncate(s, 10)
    assert len(out) == 10
    assert out.endswith("…")
