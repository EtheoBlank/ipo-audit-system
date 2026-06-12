"""综合底稿自动生成 — Streamlit 前端。

流程：
  1. 上传 Excel 模板 → 解析
  2. 选择项目 → 自动跑填充（workpaper → rule → web_search → calculated）
  3. 显示未填字段的"一次性问答"，用户回答后系统写回
  4. 预览与导出最终 Excel
"""
from __future__ import annotations

import asyncio
import io
import logging
from dataclasses import dataclass
from typing import Any, Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.comprehensive.builtin_rules import default_rule_book
from app.services.comprehensive.fill_engine import ComprehensiveFillEngine
from app.services.comprehensive.field_mapper import (
    FieldMapper,
    WorkpaperDataContext,
)
from app.services.comprehensive.qa_engine import QAEngine
from app.services.comprehensive.rule_engine import RuleEngine
from app.services.comprehensive.schemas import (
    FillReport,
    TemplateSchema,
)
from app.services.comprehensive.template_parser import TemplateParser
from app.services.comprehensive.web_search_engine import (
    SearchHit,
    WebSearchEngine,
)

logger = logging.getLogger(__name__)

API_BASE_URL = "http://localhost:8000"


# ============================================================
# Streamlit 工具
# ============================================================

def _run_async(coro):
    """在 Streamlit（同步）上下文中跑 async 协程。"""
    return asyncio.run(coro)


def _load_template_schema(uploaded_file) -> Optional[TemplateSchema]:
    """解析用户上传的模板。"""
    try:
        data = uploaded_file.read()
        return TemplateParser().parse(data)
    except Exception as exc:  # noqa: BLE001
        st.error(f"模板解析失败：{exc}")
        return None


def _build_context(project_id: int) -> WorkpaperDataContext:
    """从后端拉取项目数据，组装上下文。

    此处采用占位实现：实际项目应通过 HTTP 调用后端 API 获取 ORM 数据。
    """
    # TODO: 接入现有 /api/projects/{id}/data 接口
    return WorkpaperDataContext(
        project=type("P", (), {
            "company_name": "ACME 科技股份有限公司",
            "industry": "制造业",
            "fiscal_year": 2024,
        })(),
        account_balances=pd.DataFrame(),
        extra={"revenue": 36500.0},
    )


def _build_engine() -> ComprehensiveFillEngine:
    """构造编排器。"""

    async def _no_reg(q, k):
        return []

    async def _no_kb(q, k):
        return []

    return ComprehensiveFillEngine(
        mapper=FieldMapper(),
        rule_engine=RuleEngine(default_rule_book()),
        web_engine=WebSearchEngine(regulation_search=_no_reg, kb_search=_no_kb),
        qa_engine=QAEngine(),
    )


def _render_schema_summary(schema: TemplateSchema) -> None:
    """渲染模板解析结果概览。"""
    st.success(
        f"✅ 模板解析成功：{schema.template_name}（{schema.template_id} v{schema.version}）"
    )
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("总字段数", len(schema.fields))
    col2.metric("工作表数", len(schema.sheets))
    col3.metric("事务所", schema.firm_id)
    col4.metric("行业", schema.industry or "—")

    # 字段分类
    by_source: dict[str, int] = {}
    for f in schema.fields:
        prefix = f.source.split(":", 1)[0] if ":" in f.source else f.source
        by_source[prefix] = by_source.get(prefix, 0) + 1
    df = pd.DataFrame(
        [{"来源类型": k, "字段数": v} for k, v in by_source.items()]
    )
    st.markdown("##### 字段来源分布")
    st.dataframe(df, use_container_width=True, hide_index=True)


def _render_fill_report(report: FillReport) -> None:
    """渲染填充结果与未填项。"""
    col1, col2, col3 = st.columns(3)
    col1.metric("总字段", report.total_fields)
    col2.metric("已自动填充", report.filled, delta_color="normal")
    col3.metric("待人工补全", report.pending, delta_color="inverse")

    st.progress(
        report.filled / report.total_fields if report.total_fields else 1.0
    )

    st.markdown("##### 填充明细")
    rows = [
        {
            "字段ID": r.field_id,
            "已填值": _truncate(str(r.value), 60),
            "来源": r.source_used,
            "置信度": f"{r.confidence:.0%}",
            "引用/依据": _truncate(r.citation or "", 80),
        }
        for r in report.results
    ]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


def _render_questions(report: FillReport, key_prefix: str) -> dict[str, str]:
    """渲染未填问题，收集用户回答。返回 {question_id: answer}。"""
    if not report.open_questions:
        st.success("🎉 全部字段已自动填完，无需补充问题。")
        return {}

    st.markdown("##### 🧠 一次性问答补全")
    st.caption(
        f"共 {len(report.open_questions)} 个主题，系统已按主题合并同类问题，"
        "请逐一回答后点击「提交所有回答」。"
    )

    answers: dict[str, str] = {}
    for i, q in enumerate(report.open_questions):
        with st.expander(
            f"问题 {i+1} · 主题：{q.topic}（覆盖 {len(q.field_ids)} 个字段）",
            expanded=True,
        ):
            st.markdown(f"**{q.prompt}**")
            with st.container():
                st.caption("上下文：")
                st.code(q.context, language="text")
            answers[q.question_id] = st.text_area(
                "你的回答",
                key=f"{key_prefix}_{q.question_id}",
                height=120,
                placeholder="请用至少 100 字描述...",
            )
    return answers


def _export_to_excel(schema: TemplateSchema, report: FillReport) -> bytes:
    """把填充结果写回原模板并返回二进制流。

    设计：
    - 值写入原占位符单元格（或命名区域）
    - 来源/置信度/引用写入独立的 ``_log`` 工作表，避免与值列冲突
    - 原模板的公式、合并、格式、其它单元格内容完全保留
    """
    # 从 session_state 取原文件
    raw: bytes = st.session_state.get("__comprehensive_template_bytes__", b"")
    if not raw:
        return b""

    wb = load_workbook(filename=io.BytesIO(raw))
    by_id = {r.field_id: r for r in report.results}

    # 独立的审计轨迹工作表
    if "_log" in wb.sheetnames:
        del wb["_log"]
    log = wb.create_sheet("_log")
    log.append([
        "field_id", "sheet", "cell", "value", "source",
        "confidence", "citation", "filled_at",
    ])

    for f in schema.fields:
        if f.field_id not in by_id or by_id[f.field_id].value is None:
            continue
        if f.sheet not in wb.sheetnames:
            continue
        ws = wb[f.sheet]
        cell = ws.cell(row=f.row, column=f.column)
        # 仅在单元格是占位符/命名区域时替换，保留其它内容（公式/手工填入）
        if isinstance(cell.value, str) and (
            cell.value.startswith("{{")
            or cell.value.endswith("}}")
        ):
            cell.value = by_id[f.field_id].value
        elif f.name_range:
            cell.value = by_id[f.field_id].value
        else:
            # 既不是占位符也不是命名区域 → 不动原值
            continue
        log.append([
            f.field_id, f.sheet, f.cell_ref,
            str(by_id[f.field_id].value),
            by_id[f.field_id].source_used,
            f"{by_id[f.field_id].confidence:.2f}",
            by_id[f.field_id].citation or "",
            "",  # filled_at 占位
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _truncate(s: str, n: int) -> str:
    return s if len(s) <= n else s[: n - 1] + "…"


# ============================================================
# 主入口
# ============================================================

def show_comprehensive_workpaper():
    st.markdown("## 📑 综合底稿自动生成")
    st.caption(
        "上传事务所的 Excel 综合底稿模板 → 选项目 → 系统自动调用 "
        "「基础底稿 + 审计手册规则 + 权威信息检索 + 一次性问答」"
        "四路数据填充。模板规范参见 docs/COMPREHENSIVE_WORKPAPER_TEMPLATE_SPEC.md"
    )

    # 1) 上传模板
    st.markdown("### 第 1 步：上传综合底稿模板")
    uploaded = st.file_uploader(
        "选择 .xlsx 文件",
        type=["xlsx"],
        key="comprehensive_template",
    )
    if uploaded is None:
        st.info("请先上传模板。")
        return

    # 暂存到 session
    st.session_state["__comprehensive_template_bytes__"] = uploaded.getvalue()

    # 2) 解析
    schema = _load_template_schema(uploaded)
    if schema is None:
        return
    _render_schema_summary(schema)
    st.session_state["__comprehensive_schema__"] = schema

    # 3) 选项目并跑填充
    st.markdown("### 第 2 步：选择项目并自动填充")
    project_id = st.number_input("项目ID", min_value=1, value=1, step=1)
    if st.button("🚀 开始自动填充", type="primary"):
        with st.spinner("正在调用四路数据源..."):
            ctx = _build_context(int(project_id))
            engine = _build_engine()
            try:
                report = _run_async(engine.fill(schema, ctx))
            except Exception as exc:  # noqa: BLE001
                st.error(f"填充失败：{exc}")
                return
        st.session_state["__comprehensive_report__"] = report
        st.success(
            f"自动填充完成：{report.filled}/{report.total_fields}，"
            f"待补全 {report.pending} 项。"
        )

    # 4) 展示结果 & 问答
    report: Optional[FillReport] = st.session_state.get("__comprehensive_report__")
    if report is None:
        return

    _render_fill_report(report)
    answers = _render_questions(report, key_prefix="qa")

    # 5) 提交回答
    if report.open_questions and st.button("📥 提交所有回答", type="primary"):
        if not any(v.strip() for v in answers.values()):
            st.warning("请至少回答一个问题。")
        else:
            with st.spinner("正在合并回答到模板..."):
                engine = _build_engine()
                report = _run_async(engine.apply_qa_answers(report, answers))
                st.session_state["__comprehensive_report__"] = report
            st.success("✅ 回答已合并到报告")
            st.rerun()

    # 6) 导出
    st.markdown("### 第 3 步：导出最终 Excel")
    if st.button("💾 生成 .xlsx"):
        xlsx_bytes = _export_to_excel(schema, report)
        if xlsx_bytes:
            st.download_button(
                "⬇️ 下载综合底稿",
                data=xlsx_bytes,
                file_name=f"{schema.template_id}_filled.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.error("导出失败：未找到原始模板字节。")
